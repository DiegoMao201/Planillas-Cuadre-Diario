# ======================================================================================
# ARCHIVO: Cuadre_Diario_Caja_Final.py
# VERSIÓN: Con Reporte Gerencial por Correo (Diseño Profesional Corregido V5)
# ======================================================================================
import streamlit as st
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
import json
import pandas as pd
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import hashlib
import yagmail
import smtplib

# --- FUNCIÓN PARA VERIFICAR LA CONTRASEÑA ---
def check_password():
    """
    Muestra un formulario de login y retorna True si la contraseña es correcta.
    """
    if st.session_state.get("authenticated", False):
        return True

    st.header("🔐 Autenticación Requerida")
    st.write("Por favor, ingrese la contraseña para acceder al formulario.")

    with st.form("login"):
        password = st.text_input("Contraseña", type="password")
        submitted = st.form_submit_button("Ingresar")

        if submitted:
            hashed_input = hashlib.sha256(password.encode()).hexdigest()
            correct_hashed_password = st.secrets["credentials"]["hashed_password"]
            
            if hashed_input == correct_hashed_password:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("La contraseña es incorrecta.")
    return False

# --- 1. CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(layout="wide", page_title="Cuadre Diario de Caja")

# --- 2. CONEXIÓN SEGURA A GOOGLE SHEETS ---
@st.cache_resource(ttl=600)
def connect_to_gsheet():
    """
    Establece conexión con Google Sheets usando las credenciales de st.secrets.
    Retorna los objetos de las hojas de trabajo necesarias.
    """
    try:
        creds_json = dict(st.secrets["google_credentials"])
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_json, scope)
        client = gspread.authorize(creds)
        sheet = client.open(st.secrets["google_sheets"]["spreadsheet_name"])
        registros_ws = sheet.worksheet(st.secrets["google_sheets"]["registros_sheet_name"])
        config_ws = sheet.worksheet(st.secrets["google_sheets"]["config_sheet_name"])
        consecutivos_ws = sheet.worksheet("Consecutivos")
        
        global_consecutivo_ws = sheet.worksheet("GlobalConsecutivo")
        
        return registros_ws, config_ws, consecutivos_ws, global_consecutivo_ws
    except Exception as e:
        st.error(f"Error fatal al conectar con Google Sheets: {e}")
        st.warning("Verifique las credenciales y los nombres de las hojas (incluyendo 'GlobalConsecutivo') en los 'secrets' de Streamlit.")
        return None, None, None, None

# --- 3. LÓGICA DE DATOS Y PROCESAMIENTO ---
def get_app_config(config_ws):
    """
    Carga la configuración esencial (tiendas, bancos, terceros) desde la hoja 'Configuracion'.
    Usa .strip() para eliminar espacios en blanco y asegura ignorar filas vacías.
    """
    try:
        config_data = config_ws.get_all_records()
        tiendas = sorted(list(set(str(d['Detalle']).strip() for d in config_data if d.get('Tipo Movimiento') == 'TIENDA' and d.get('Detalle'))))
        bancos = sorted(list(set(str(d['Detalle']).strip() for d in config_data if d.get('Tipo Movimiento') == 'BANCO' and d.get('Detalle'))))
        terceros = sorted(list(set(str(d['Detalle']).strip() for d in config_data if d.get('Tipo Movimiento') == 'TERCERO' and d.get('Detalle'))))
        return tiendas, bancos, terceros
    except Exception as e:
        st.error(f"Error al cargar la configuración de tiendas, bancos y terceros: {e}")
        return [], [], []

def get_account_mappings(config_ws):
    """
    Crea un diccionario de mapeo de cuentas a partir de la hoja 'Configuracion'.
    """
    try:
        records = config_ws.get_all_records()
        mappings = {}
        for record in records:
            tipo = record.get("Tipo Movimiento")
            detalle = record.get("Detalle")
            cuenta = record.get("Cuenta Contable")

            if detalle and cuenta:
                detalle_str = str(detalle).strip()
                cuenta_str = str(cuenta)
                if tipo in ["BANCO", "TERCERO"]:
                    mappings[detalle_str] = {
                        'cuenta': cuenta_str,
                        'nit': str(record.get("NIT", "")),
                        'nombre': str(record.get("Nombre Tercero", ""))
                    }
                elif tipo in ["GASTO", "TARJETA", "EFECTIVO"]:
                    mappings[detalle_str] = {'cuenta': cuenta_str}
        return mappings
    except Exception as e:
        st.error(f"Error al leer el mapeo de cuentas. Revisa la estructura de la hoja 'Configuracion'. Error: {e}")
        return {}

def generate_txt_file(registros_ws, config_ws, start_date, end_date, selected_store):
    """
    Genera el contenido del archivo TXT para el ERP, con filtros por fecha y tienda.
    """
    st.info("Generando archivo TXT... Esto puede tardar unos segundos.")
    
    all_records = registros_ws.get_all_records()
    account_mappings = get_account_mappings(config_ws)

    if not account_mappings:
        st.error("No se pudo generar el reporte: Faltan mapeos de cuentas en 'Configuracion'.")
        return None

    try:
        date_filtered_records = [
            r for r in all_records
            if start_date <= datetime.strptime(r.get('Fecha', '01/01/1900'), '%d/%m/%Y').date() <= end_date
        ]
        if selected_store == "Todas las Tiendas":
            filtered_records = date_filtered_records
        else:
            filtered_records = [r for r in date_filtered_records if str(r.get('Tienda', '')).strip() == selected_store]

    except ValueError as e:
        st.error(f"Error de formato de fecha en 'Registros'. Asegúrese que las fechas sean DD/MM/YYYY. Error: {e}")
        return None

    if not filtered_records:
        st.warning("No se encontraron registros en el rango de fechas y tienda seleccionados.")
        return None

    filtered_records.sort(key=lambda r: (r.get('Tienda', ''), r.get('Fecha', '')))
    txt_lines = []
    
    for record in filtered_records:
        consecutivo_referencia = record.get('Consecutivo_Asignado', '0')
        consecutivo_documento = record.get('Consecutivo_Global_Doc', '0')
        
        tienda = str(record.get('Tienda', ''))
        fecha_cuadre = record['Fecha']
        centro_costo = tienda 
        tienda_descripcion = re.sub(r'[\(\)]', '', tienda).strip()
        total_debito_dia = 0

        movimientos = {
            'TARJETA': json.loads(record.get('Tarjetas', '[]')),
            'CONSIGNACION': json.loads(record.get('Consignaciones', '[]')),
            'GASTO': json.loads(record.get('Gastos', '[]')),
            'EFECTIVO': json.loads(record.get('Efectivo', '[]'))
        }

        for tipo_mov, data_list in movimientos.items():
            for item in data_list:
                valor = float(item.get('Valor', 0))
                if valor == 0: continue
                total_debito_dia += valor

                cuenta = ""
                nit_tercero = "0"
                nombre_tercero_final = "0" 
                serie_documento = centro_costo
                descripcion = f"Ventas planillas contado {tienda_descripcion}"
                
                if tipo_mov == 'TARJETA':
                    cuenta = account_mappings.get('Tarjetas', {}).get('cuenta', 'ERR_TARJETA')
                    serie_documento = f"T{centro_costo}"
                    fecha_tarjeta = item.get('Fecha', '')
                    descripcion = f"Ventas planillas contado Tarjeta {fecha_tarjeta} - {tienda_descripcion}"

                elif tipo_mov == 'CONSIGNACION':
                    banco = item.get('Banco')
                    cuenta = account_mappings.get(banco, {}).get('cuenta', f'ERR_{banco}')
                    fecha_consignacion = item.get('Fecha', '')
                    descripcion = f"Ventas planillas contado consignacion {fecha_consignacion} - {tienda_descripcion}"

                elif tipo_mov == 'GASTO':
                    gasto_tercero = item.get('Tercero')
                    
                    if gasto_tercero and gasto_tercero != "N/A":
                        tercero_info = account_mappings.get(gasto_tercero)
                        if tercero_info:
                            cuenta = tercero_info.get('cuenta', f'ERR_TERCERO_{gasto_tercero}')
                            nit_tercero = tercero_info.get('nit', '0')
                            nombre_tercero_desc = tercero_info.get('nombre', gasto_tercero)
                            descripcion = f"{item.get('Descripción', 'Gasto')} - {nombre_tercero_desc}"
                        else:
                            cuenta = account_mappings.get('Reintegro Caja Menor', {}).get('cuenta', 'ERR_GASTO')
                            descripcion = f"{item.get('Descripción', 'Gasto')} (Tercero {gasto_tercero} no encontrado)"
                    else:
                        cuenta = account_mappings.get('Reintegro Caja Menor', {}).get('cuenta', 'ERR_GASTO')
                        descripcion = item.get('Descripción', 'Gasto Varios')

                elif tipo_mov == 'EFECTIVO':
                    tipo_especifico = item.get('Tipo', 'Efectivo Entregado')
                    destino_tercero = item.get('Destino/Tercero (Opcional)')
                    
                    if tipo_especifico == "Efectivo Entregado" and destino_tercero and destino_tercero != "N/A":
                        tercero_info = account_mappings.get(destino_tercero)
                        if tercero_info:
                            cuenta = tercero_info.get('cuenta', f'ERR_TERCERO_{destino_tercero}')
                            nit_tercero = tercero_info.get('nit', '0')
                            nombre_tercero_desc = tercero_info.get('nombre', destino_tercero)
                            descripcion = f"Ventas planillas contado Entrega efectivo a {nombre_tercero_desc} - {tienda_descripcion}"
                        else:
                            cuenta = account_mappings.get(tipo_especifico, {}).get('cuenta', f'ERR_{tipo_especifico}')
                            descripcion = f"Ventas planillas contado Entrega efectivo a TERCERO_NO_ENCONTRADO({destino_tercero}) - {tienda_descripcion}"
                    else:
                        cuenta = account_mappings.get(tipo_especifico, {}).get('cuenta', f'ERR_{tipo_especifico}')

                linea = "|".join([
                    fecha_cuadre, str(consecutivo_documento), str(cuenta), "8",
                    descripcion, serie_documento, str(consecutivo_referencia),
                    str(valor), "0", centro_costo, nit_tercero, nombre_tercero_final, "0"
                ])
                txt_lines.append(linea)

        # Línea de contrapartida (crédito)
        if total_debito_dia > 0:
            cuenta_venta = "11050501"
            descripcion_credito = f"Ventas planillas contado {tienda_descripcion}"
            
            linea_credito = "|".join([
                fecha_cuadre, str(consecutivo_documento), str(cuenta_venta), "8",
                descripcion_credito, centro_costo, str(consecutivo_referencia),
                "0", str(total_debito_dia), centro_costo, "0", "0", "0"
            ])
            txt_lines.append(linea_credito)
            
    return "\n".join(txt_lines)

# --- GENERADOR DE EXCEL ---
def generate_excel_report(registros_ws, start_date, end_date, selected_store):
    """
    Genera un archivo Excel profesional y detallado para la revisión del cuadre de caja.
    """
    st.info("Generando reporte Excel profesional... Esto puede tardar unos segundos.")

    try:
        all_records = registros_ws.get_all_records()
        date_filtered_records = [
            r for r in all_records
            if start_date <= datetime.strptime(r.get('Fecha', '01/01/1900'), '%d/%m/%Y').date() <= end_date
        ]
        if selected_store == "Todas las Tiendas":
            filtered_records = date_filtered_records
        else:
            filtered_records = [r for r in date_filtered_records if str(r.get('Tienda', '')).strip() == selected_store]
    except Exception as e:
        st.error(f"Error al filtrar registros para Excel: {e}")
        return None

    if not filtered_records:
        st.warning("No se encontraron registros para generar el reporte Excel.")
        return None
    
    filtered_records.sort(key=lambda r: (datetime.strptime(r.get('Fecha', '01/01/1900'), '%d/%m/%Y'), r.get('Tienda', '')))

    output = io.BytesIO()
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Reporte Cuadre de Caja"

    # Estilos
    font_title = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
    fill_title = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    font_header = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font_day_header = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    fill_day_header = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    font_category = Font(name='Calibri', size=11, bold=True)
    fill_category = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    font_total_label = Font(name='Calibri', size=12, bold=True)
    font_total_value = Font(name='Calibri', size=12, bold=True)
    fill_summary = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    font_diff_ok = Font(name='Calibri', size=12, bold=True, color="00B050")
    font_diff_bad = Font(name='Calibri', size=12, bold=True, color="C00000")
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    currency_format = '$ #,##0'

    # Título Principal
    ws.merge_cells('A1:F2')
    title_cell = ws['A1']
    title_cell.value = f"REPORTE DE CUADRE DIARIO - {selected_store.upper()}"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = align_center

    ws.merge_cells('A3:F3')
    ws['A3'].value = f"Período del {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}"
    ws['A3'].alignment = align_center
    ws['A3'].font = Font(name='Calibri', size=12, italic=True)
    
    current_row = 5

    # Iterar sobre cada registro
    for record in filtered_records:
        fecha_str = record.get('Fecha', 'N/A')
        tienda = record.get('Tienda', 'N/A')
        
        # Cabecera del Día
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        day_header_cell = ws.cell(row=current_row, column=1, value=f"Resumen del Día: {fecha_str} - Tienda: {tienda}")
        day_header_cell.font = font_day_header
        day_header_cell.fill = fill_day_header
        day_header_cell.alignment = align_center
        current_row += 1

        # Cabeceras de la tabla
        headers = ["Tipo de Movimiento", "Fecha Específica", "Detalle", "Tercero / Banco", "Valor"]
        for col_num, header_title in enumerate(headers, 2):
            cell = ws.cell(row=current_row, column=col_num, value=header_title)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = thin_border
            cell.alignment = align_center
        current_row += 1
        
        # Procesar movimientos
        total_desglose = 0
        subtotales = {'Tarjetas': 0, 'Consignaciones': 0, 'Gastos': 0, 'Efectivo': 0}
        
        movimientos_map = {
            'Tarjetas': ('Tarjetas', '[]'),
            'Consignaciones': ('Consignaciones', '[]'),
            'Gastos': ('Gastos', '[]'),
            'Efectivo': ('Efectivo', '[]')
        }

        for cat_name, (json_key, default_val) in movimientos_map.items():
            data_list = json.loads(record.get(json_key, default_val))
            if not data_list: continue

            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
            cat_cell = ws.cell(row=current_row, column=2, value=cat_name.upper())
            cat_cell.font = font_category
            cat_cell.fill = fill_category
            cat_cell.border = thin_border
            current_row += 1

            for item in data_list:
                valor = float(item.get('Valor', 0))
                if valor == 0: continue
                
                ws.cell(row=current_row, column=2, value=item.get('Tipo', cat_name.rstrip('s')))
                ws.cell(row=current_row, column=3, value=item.get('Fecha', fecha_str))
                ws.cell(row=current_row, column=4, value=item.get('Descripción', 'N/A'))
                ws.cell(row=current_row, column=5, value=item.get('Tercero', item.get('Banco', item.get('Destino/Tercero (Opcional)', 'N/A'))))
                valor_cell = ws.cell(row=current_row, column=6, value=valor)
                valor_cell.number_format = currency_format
                valor_cell.alignment = align_right

                for col_idx in range(2, 7):
                    ws.cell(row=current_row, column=col_idx).border = thin_border

                total_desglose += valor
                subtotales[cat_name] += valor
                current_row += 1
        
        # Bloque de Resumen
        current_row += 1
        
        venta_total_sistema = float(record.get('Venta_Total_Dia', 0))
        diferencia = venta_total_sistema - total_desglose

        summary_data = [
            ("Venta Total (Sistema)", venta_total_sistema),
            ("Total Tarjetas", subtotales['Tarjetas']),
            ("Total Consignaciones", subtotales['Consignaciones']),
            ("Total Gastos", subtotales['Gastos']),
            ("Total Efectivo (Entregas/Reintegros)", subtotales['Efectivo']),
            ("TOTAL DESGLOSADO (Suma de Movimientos)", total_desglose),
            ("DIFERENCIA EN CUADRE", diferencia)
        ]

        for label, value in summary_data:
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
            label_cell = ws.cell(row=current_row, column=4, value=label)
            value_cell = ws.cell(row=current_row, column=6, value=value)
            
            label_cell.font = font_total_label
            label_cell.alignment = align_right
            label_cell.fill = fill_summary
            label_cell.border = thin_border
            ws.cell(row=current_row, column=5).border = thin_border
            
            value_cell.font = font_total_value
            value_cell.number_format = currency_format
            value_cell.alignment = align_right
            value_cell.fill = fill_summary
            value_cell.border = thin_border

            if "DIFERENCIA" in label:
                if diferencia == 0:
                    value_cell.font = font_diff_ok
                else:
                    value_cell.font = font_diff_bad

            current_row += 1

        current_row += 2

    # Ajustar Ancho de Columnas
    column_widths = {'A': 5, 'B': 22, 'C': 18, 'D': 35, 'E': 25, 'F': 18}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    workbook.save(output)
    output.seek(0)
    
    return output.getvalue()

# ======================================================================================
# --- INICIO DE LA SECCIÓN DE CORREO ELECTRÓNICO (NUEVO DISEÑO MEJORADO) ---
# ======================================================================================

def format_cop(value):
    """
    Formatea un número como moneda colombiana (ej: $ 1.234.500)
    sin depender del módulo 'locale' para evitar errores en el servidor.
    """
    if not isinstance(value, (int, float)):
        return "$ 0"
    formatted_value = f"{value:,.0f}"
    formatted_value = formatted_value.replace(",", ".")
    return f"$ {formatted_value}"

def generate_professional_email_body(records, start_date, end_date, selected_store):
    """
    Genera un cuerpo de correo HTML profesional y visualmente atractivo.
    El diseño es más compacto y muestra los totales por categoría.
    """
    if start_date == end_date:
        report_date_str = f"Reporte del día: {start_date.strftime('%d/%m/%Y')}"
        subtitle_text = f"Reporte para: {selected_store}"
    else:
        report_date_str = f"Período del {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}"
        subtitle_text = f"Reporte Consolidado para: {selected_store}"
    
    report_time_str = datetime.now().strftime("%d/%m/%Y a las %H:%M")
    
    # Agrupar los registros por Tienda y luego por Fecha
    grouped_records = {}
    for record in records:
        tienda = record.get('Tienda', 'Desconocida')
        if tienda not in grouped_records:
            grouped_records[tienda] = []
        grouped_records[tienda].append(record)
    
    # Construcción de las tablas de resumen detalladas
    details_html = ""
    for tienda, records_list in grouped_records.items():
        details_html += f"""
        <table border="0" cellpadding="0" cellspacing="0" width="100%" style="margin-top: 20px;">
            <tr>
                <td align="left" style="background-color: #0077b6; padding: 10px; border-radius: 8px 8px 0 0;">
                    <h3 style="color: #ffffff; font-size: 16px; margin: 0; padding-left: 5px;">Tienda: {tienda}</h3>
                </td>
            </tr>
            <tr>
                <td style="border: 1px solid #e9ecef; border-top: none; padding: 16px; border-radius: 0 0 8px 8px;">
                    <table border="0" cellpadding="0" cellspacing="0" width="100%">
        """
        
        # Resumen por fecha dentro de la tienda
        date_grouped_records = {}
        for r in records_list:
            fecha = r.get('Fecha', 'Desconocida')
            if fecha not in date_grouped_records:
                date_grouped_records[fecha] = []
            date_grouped_records[fecha].append(r)
            
        for fecha, daily_records in date_grouped_records.items():
            venta_total_sistema = sum(float(r.get('Venta_Total_Dia', 0)) for r in daily_records)
            
            total_tarjetas = sum(sum(float(t.get('Valor', 0)) for t in json.loads(r.get('Tarjetas', '[]'))) for r in daily_records)
            total_consignaciones = sum(sum(float(c.get('Valor', 0)) for c in json.loads(r.get('Consignaciones', '[]'))) for r in daily_records)
            total_gastos = sum(sum(float(g.get('Valor', 0)) for g in json.loads(r.get('Gastos', '[]'))) for r in daily_records)
            total_efectivo = sum(sum(float(e.get('Valor', 0)) for e in json.loads(r.get('Efectivo', '[]'))) for r in daily_records)
            
            total_desglose = total_tarjetas + total_consignaciones + total_gastos + total_efectivo
            diferencia = venta_total_sistema - total_desglose
            
            # Subcard dentro del detalle de movimientos
            details_html += f"""
            <tr>
                <td colspan="2">
                    <div style="background-color:#fff; border-radius:10px; border:1px solid #e0e0e0; padding:15px; margin-top:15px;">
                        <h4 style="font-size:16px; font-weight:600; margin:0 0 10px 0; color:#333;">Resumen del día: {fecha}</h4>
                        <table border="0" cellpadding="0" cellspacing="0" width="100%">
                            <tr>
                                <td style="padding:5px 0; font-size:14px; color:#555;">💰 Venta Total</td>
                                <td align="right" style="padding:5px 0; font-size:14px; font-weight:bold; color:#007200;">{format_cop(venta_total_sistema)}</td>
                            </tr>
                            <tr>
                                <td style="padding:5px 0; font-size:14px; color:#555;">💳 Total Tarjetas</td>
                                <td align="right" style="padding:5px 0; font-size:14px; font-weight:bold;">{format_cop(total_tarjetas)}</td>
                            </tr>
                            <tr>
                                <td style="padding:5px 0; font-size:14px; color:#555;">🏦 Total Consignaciones</td>
                                <td align="right" style="padding:5px 0; font-size:14px; font-weight:bold;">{format_cop(total_consignaciones)}</td>
                            </tr>
                            <tr>
                                <td style="padding:5px 0; font-size:14px; color:#555;">💸 Total Gastos</td>
                                <td align="right" style="padding:5px 0; font-size:14px; font-weight:bold;">{format_cop(total_gastos)}</td>
                            </tr>
                            <tr>
                                <td style="padding:5px 0; font-size:14px; color:#555;">💵 Total Efectivo</td>
                                <td align="right" style="padding:5px 0; font-size:14px; font-weight:bold;">{format_cop(total_efectivo)}</td>
                            </tr>
                            <tr>
                                <td style="padding:5px 0; font-size:14px; color:#555; border-top:1px dashed #ccc;">Total Desglose</td>
                                <td align="right" style="padding:5px 0; font-size:14px; font-weight:bold; color:#023e8a; border-top:1px dashed #ccc;">{format_cop(total_desglose)}</td>
                            </tr>
                            <tr>
                                <td style="padding:5px 0; font-size:14px; color:#555; border-top:1px dashed #ccc;">Diferencia</td>
                                <td align="right" style="padding:5px 0; font-size:14px; font-weight:bold; color:{'#00B050' if diferencia == 0 else '#C00000'}; border-top:1px dashed #ccc;">{format_cop(diferencia)}</td>
                            </tr>
                        </table>
                    </div>
                </td>
            </tr>
            """
        details_html += """
                    </table>
                </td>
            </tr>
        </table>
        """

    # HTML principal del correo con el diseño compacto y corregido
    html_body = f"""<!DOCTYPE html>
<html lang="es" xmlns="http://www.w3.org/1999/xhtml" xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Resumen Gerencial de Caja</title>
    <style type="text/css">
        @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@400;500;700&display=swap');
        body, table, td, a {{ -webkit-text-size-adjust: 100%; -ms-text-size-adjust: 100%; font-family: 'Roboto', sans-serif; }}
        table, td {{ mso-table-lspace: 0pt; mso-table-rspace: 0pt; }}
        img {{ -ms-interpolation-mode: bicubic; border: 0; height: auto; line-height: 100%; outline: none; text-decoration: none; }}
        body {{ height: 100% !important; margin: 0 !important; padding: 0 !important; width: 100% !important; background-color: #f8f9fa; }}
    </style>
</head>
<body style="margin: 0; padding: 0; background-color: #f8f9fa;">
    <center style="width: 100%; background-color: #f8f9fa;">
        <div style="background-color:#f8f9fa; max-width:600px; margin:0 auto; padding: 20px 0;">
            <table border="0" cellpadding="0" cellspacing="0" width="100%" style="max-width: 600px; background-color: #ffffff; border-radius: 12px; border: 1px solid #dee2e6; overflow: hidden; box-shadow: 0 4px 8px rgba(0,0,0,0.05);">
                <tr>
                    <td align="center" style="background: linear-gradient(90deg, #023e8a, #0077b6); padding: 25px; border-radius: 12px 12px 0 0;">
                        <h1 style="font-size: 24px; font-weight: 700; margin: 0 0 5px 0; color: #ffffff;">Resumen Gerencial de Caja</h1>
                        <p style="font-size: 14px; font-weight: 400; color: #caf0f8; margin: 0;">{subtitle_text}</p>
                        <div style="margin-top: 15px;">
                            <span style="font-size: 13px; font-weight: 500; color: #ffffff; text-decoration: none; padding: 6px 15px; border-radius: 15px; display: inline-block; background-color: rgba(255, 255, 255, 0.15);">{report_date_str}</span>
                        </div>
                    </td>
                </tr>
                <tr>
                    <td align="left" style="padding: 24px;">
                        <h2 style="color: #023e8a; font-size: 18px; font-weight: 700; margin: 0 0 16px 0; padding-bottom: 10px; border-bottom: 2px solid #0077b6;">Detalle de Movimientos</h2>
                        {details_html}
                    </td>
                </tr>
                <tr>
                    <td bgcolor="#f1f3f5" align="center" style="padding: 16px; border-radius: 0 0 12px 12px;">
                        <p style="font-size: 12px; color: #6c757d; margin: 0;">Este reporte fue generado automáticamente el {report_time_str}.<br><strong style="color: #495057;">Sistema de Gestión Financiera</strong></p>
                    </td>
                </tr>
            </table>
        </div>
    </center>
</body>
</html>"""
    return html_body

def send_summary_email(registros_ws, start_date, end_date, selected_store, recipient_email):
    """
    Filtra los datos, genera el resumen con el nuevo diseño y lo envía por correo.
    """
    st.info("Preparando y enviando resumen gerencial...")

    try:
        sender_email = st.secrets["email_credentials"]["sender_email"]
        sender_password = st.secrets["email_credentials"]["sender_password"]
    except (KeyError, TypeError):
        st.error("Credenciales de correo no encontradas o mal configuradas en los 'secrets' de Streamlit.")
        st.warning("Asegúrese de tener una sección [email_credentials] con 'sender_email' y 'sender_password'.")
        return

    try:
        all_records = registros_ws.get_all_records()
        date_filtered_records = [
            r for r in all_records
            if start_date <= datetime.strptime(r.get('Fecha', '01/01/1900'), '%d/%m/%Y').date() <= end_date
        ]
        if selected_store == "Todas las Tiendas":
            filtered_records = date_filtered_records
        else:
            filtered_records = [r for r in date_filtered_records if str(r.get('Tienda', '')).strip() == selected_store]
    except Exception as e:
        st.error(f"Error al filtrar los registros para el correo: {e}")
        return

    if not filtered_records:
        st.warning("No se encontraron registros en el rango de fechas y tienda seleccionados para enviar el correo.")
        return

    # Llamada a la NUEVA función para generar el cuerpo del correo
    email_body = generate_professional_email_body(filtered_records, start_date, end_date, selected_store)
    subject = f"Resumen de Cuadre de Caja - {selected_store} - {start_date.strftime('%d/%m')} a {end_date.strftime('%d/%m')}"

    try:
        yag = yagmail.SMTP(sender_email, sender_password)
        yag.send(
            to=recipient_email,
            subject=subject,
            contents=email_body
        )
    except smtplib.SMTPAuthenticationError:
        st.error("Error de autenticación con el servidor de correo. Verifique el email y la contraseña de aplicación en los 'secrets'.")
        return
    except Exception as e:
        st.error(f"Ocurrió un error inesperado al enviar el correo: {e}")
        return
    
    st.success(f"¡Resumen gerencial enviado exitosamente a {recipient_email}!")

# ======================================================================================
# --- FIN DE LA SECCIÓN DE CORREO ELECTRÓNICO ---
# ======================================================================================

# --- 4. GESTIÓN DEL ESTADO DE LA SESIÓN ---
def initialize_session_state():
    """Inicializa el estado de la sesión para almacenar datos del formulario."""
    defaults = {
        'page': 'Formulario', 'venta_total_dia': 0.0, 'factura_inicial': "", 'factura_final': "",
        'tarjetas': [], 'consignaciones': [], 'gastos': [], 'efectivo': [],
        'authenticated': False
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

def clear_form_state():
    """Limpia el formulario, conservando la tienda, fecha y estado de autenticación."""
    tienda = st.session_state.get('tienda_seleccionada', None)
    fecha = st.session_state.get('fecha_seleccionada', datetime.now().date())
    auth_status = st.session_state.get('authenticated', False)
    
    keys_to_keep = ['page', 'tienda_seleccionada', 'fecha_seleccionada', 'authenticated']
    for key in list(st.session_state.keys()):
        if key not in keys_to_keep:
            del st.session_state[key]
            
    initialize_session_state()
    st.session_state.tienda_seleccionada = tienda
    st.session_state.fecha_seleccionada = fecha
    st.session_state.authenticated = auth_status

# --- 5. COMPONENTES DE LA INTERFAZ DE USUARIO ---
def format_currency(num):
    """Formatea un número como moneda colombiana (ej: $1.234.567)."""
    return f"${int(num):,}".replace(",", ".") if isinstance(num, (int, float)) else "$0"

def load_cuadre_data(registros_ws):
    """Carga los datos de un cuadre existente desde la hoja 'Registros'."""
    if not st.session_state.get("tienda_seleccionada"):
        st.warning("Por favor, seleccione una tienda primero.")
        return

    id_registro = f"{st.session_state.tienda_seleccionada}-{st.session_state.fecha_seleccionada.strftime('%d/%m/%Y')}"
    try:
        cell = registros_ws.find(id_registro, in_column=1)
        if cell:
            row_data = registros_ws.row_values(cell.row)
            clear_form_state()

            st.session_state.factura_inicial = row_data[4] if len(row_data) > 4 else ""
            st.session_state.factura_final = row_data[5] if len(row_data) > 5 else ""
            st.session_state.venta_total_dia = float(row_data[6]) if len(row_data) > 6 and row_data[6] else 0.0
            st.session_state.tarjetas = json.loads(row_data[7]) if len(row_data) > 7 and row_data[7] else []
            st.session_state.consignaciones = json.loads(row_data[8]) if len(row_data) > 8 and row_data[8] else []
            st.session_state.gastos = json.loads(row_data[9]) if len(row_data) > 9 and row_data[9] else []
            st.session_state.efectivo = json.loads(row_data[10]) if len(row_data) > 10 and row_data[10] else []
            st.toast(f"✅ Cuadre para '{st.session_state.tienda_seleccionada}' cargado.", icon="📄")
        else:
            st.warning("No se encontró un cuadre para esta tienda y fecha. Puede crear uno nuevo.")
            clear_form_state()
    except Exception as e:
        st.error(f"Error al cargar datos. Verifique la hoja 'Registros'. Error: {e}")
        clear_form_state()

# --- FUNCIONES PARA MANEJAR CONSECUTIVOS ---

def get_next_consecutive(consecutivos_ws, tienda):
    """Obtiene el siguiente número consecutivo para una tienda."""
    try:
        cell = consecutivos_ws.find(tienda, in_column=1)
        if cell:
            last_consecutive = int(consecutivos_ws.cell(cell.row, 2).value)
            return last_consecutive + 1
        else:
            st.warning(f"No se encontró consecutivo para '{tienda}'. Se usará '1000' por defecto.")
            return 1000
    except Exception as e:
        st.error(f"Error al obtener consecutivo de tienda: {e}")
        return None

def update_consecutive(consecutivos_ws, tienda, new_consecutive):
    """Actualiza el último consecutivo usado para una tienda."""
    try:
        cell = consecutivos_ws.find(tienda, in_column=1)
        if cell:
            consecutivos_ws.update_cell(cell.row, 2, new_consecutive)
        else:
            consecutivos_ws.append_row([tienda, new_consecutive])
    except Exception as e:
        st.error(f"Error al actualizar consecutivo de tienda: {e}")

def get_next_global_consecutive(global_consecutivo_ws):
    """Obtiene el siguiente número consecutivo global."""
    try:
        last_consecutive = int(global_consecutivo_ws.acell('B1').value)
        return last_consecutive + 1
    except Exception as e:
        st.error(f"Error al obtener consecutivo global desde la hoja 'GlobalConsecutivo': {e}")
        st.warning("Asegúrese que la hoja exista y que la celda B1 contenga un número.")
        return None

def update_global_consecutive(global_consecutivo_ws, new_consecutive):
    """Actualiza el último consecutivo global usado."""
    try:
        global_consecutivo_ws.update_acell('B1', new_consecutive)
    except Exception as e:
        st.error(f"Error al actualizar el consecutivo global: {e}")

def display_dynamic_list_section(title, key, form_inputs, options_map=None):
    """Función reutilizable para crear secciones del formulario."""
    if options_map is None: options_map = {}

    with st.expander(f"**{title}**", expanded=True):
        with st.form(f"form_{key}", clear_on_submit=True):
            cols = st.columns(len(form_inputs))
            data = {}
            for i, (input_key, input_type, options) in enumerate(form_inputs):
                label = options.get('label', input_key)
                if input_type == "selectbox":
                    data[input_key] = cols[i].selectbox(label, options=options_map.get(input_key, []), label_visibility="collapsed", placeholder=label)
                elif input_type == "number_input":
                    data[input_key] = cols[i].number_input(label, min_value=0.0, step=1000.0, format="%.0f", label_visibility="collapsed", placeholder=label)
                elif input_type == "date_input":
                    data[input_key] = cols[i].date_input(label, value=datetime.now().date(), label_visibility="collapsed", format="DD/MM/YYYY")
                else:
                    data[input_key] = cols[i].text_input(label, label_visibility="collapsed", placeholder=label)
            
            if st.form_submit_button(f"✚ Agregar {title.split(' ')[1]}", use_container_width=True):
                if data.get("Valor", 0) > 0:
                    if 'Fecha' in data and hasattr(data['Fecha'], 'strftime'):
                        data['Fecha'] = data['Fecha'].strftime("%d/%m/%Y")
                    st.session_state[key].append(data)
                    st.toast(f"✅ {title.split(' ')[1]} agregado.")
                    st.rerun()
                else:
                    st.warning("El valor debe ser mayor a cero.")

        if st.session_state[key]:
            df = pd.DataFrame(st.session_state[key])
            df['Eliminar'] = False
            column_config = {
                "Valor": st.column_config.NumberColumn("Valor", format="$ %.0f", required=True),
                "Eliminar": st.column_config.CheckboxColumn("Eliminar", width="small")
            }
            for col_name, options_list in options_map.items():
                if col_name in df.columns:
                    column_config[col_name] = st.column_config.SelectboxColumn(col_name, options=options_list, required=True)

            edited_df = st.data_editor(df, key=f'editor_{key}', hide_index=True, use_container_width=True, column_config=column_config)

            if edited_df['Eliminar'].any():
                indices_to_remove = edited_df[edited_df['Eliminar']].index
                st.session_state[key] = [item for i, item in enumerate(st.session_state[key]) if i not in indices_to_remove]
                st.toast("🗑️ Registro(s) eliminado(s).")
                st.rerun()
            else:
                st.session_state[key] = edited_df.drop(columns=['Eliminar']).to_dict('records')

        subtotal = sum(float(item.get('Valor', 0)) for item in st.session_state[key])
        st.metric(f"Subtotal {title.split(' ')[1]}", format_currency(subtotal))

def display_tarjetas_section():
    """Muestra la sección para agregar y editar pagos con tarjeta."""
    with st.expander("💳 **Tarjetas**", expanded=True):
        with st.form("form_tarjetas", clear_on_submit=True):
            c1, c2 = st.columns(2)
            valor = c1.number_input("Valor", min_value=1.0, step=1000.0, format="%.0f", label_visibility="collapsed", placeholder="Valor Tarjeta")
            fecha = c2.date_input("Fecha", value=datetime.now().date(), label_visibility="collapsed", format="DD/MM/YYYY")
            
            if st.form_submit_button("✚ Agregar Tarjeta", use_container_width=True):
                if valor > 0:
                    st.session_state.tarjetas.append({
                        'Valor': valor,
                        'Fecha': fecha.strftime("%d/%m/%Y")
                    })
                    st.toast(f"Agregado: {format_currency(valor)}")
                    st.rerun()

        if st.session_state.tarjetas:
            df = pd.DataFrame(st.session_state.tarjetas)
            df['Eliminar'] = False
            
            if 'Fecha' in df.columns:
                df = df[['Fecha', 'Valor', 'Eliminar']]

            edited_df = st.data_editor(
                df, key='editor_tarjetas', hide_index=True, use_container_width=True,
                column_config={
                    "Valor": st.column_config.NumberColumn("Valor", format="$ %.0f", required=True),
                    "Fecha": st.column_config.TextColumn("Fecha", required=True),
                    "Eliminar": st.column_config.CheckboxColumn("Eliminar", width="small")
                }
            )
            
            if edited_df['Eliminar'].any():
                st.session_state.tarjetas = [t for i, t in enumerate(st.session_state.tarjetas) if i not in edited_df[edited_df['Eliminar']].index]
                st.toast("Tarjeta(s) eliminada(s).")
                st.rerun()
            else:
                st.session_state.tarjetas = edited_df.drop(columns=['Eliminar']).to_dict('records')
                
        st.metric("Subtotal Tarjetas", format_currency(sum(float(t.get('Valor', 0)) for t in st.session_state.tarjetas)))

def display_consignaciones_section(bancos_list):
    display_dynamic_list_section(
        "🏦 Consignaciones", "consignaciones",
        [("Banco", "selectbox", {"label": "Banco"}),
         ("Valor", "number_input", {"label": "Valor"}),
         ("Fecha", "date_input", {"label": "Fecha"})],
        options_map={"Banco": bancos_list}
    )

def display_gastos_section(terceros_list):
    terceros_con_na = ["N/A"] + terceros_list
    display_dynamic_list_section(
        "💸 Gastos", "gastos",
        [("Descripción", "text_input", {"label": "Descripción del Gasto"}),
         ("Tercero", "selectbox", {"label": "Proveedor (Opcional)"}),
         ("Valor", "number_input", {"label": "Valor"})],
        options_map={"Tercero": terceros_con_na}
    )

def display_efectivo_section(terceros_list):
    terceros_con_na = ["N/A"] + terceros_list
    display_dynamic_list_section(
        "💵 Efectivo", "efectivo",
        [("Tipo", "selectbox", {"label": "Tipo de Movimiento"}),
         ("Destino/Tercero (Opcional)", "selectbox", {"label": "Proveedor / Destino"}),
         ("Valor", "number_input", {"label": "Valor"})],
        options_map={
            "Tipo": ["Efectivo Entregado", "Reintegro Caja Menor"],
            "Destino/Tercero (Opcional)": terceros_con_na
        }
    )

def display_summary_and_save(worksheets):
    st.header("3. Verificación y Guardado", anchor=False, divider="rainbow")
    
    registros_ws, _, consecutivos_ws, global_consecutivo_ws = worksheets

    with st.container(border=True):
        sub_t = sum(float(t.get('Valor', 0)) for t in st.session_state.tarjetas)
        sub_c = sum(float(c.get('Valor', 0)) for c in st.session_state.consignaciones)
        sub_g = sum(float(g.get('Valor', 0)) for g in st.session_state.gastos)
        sub_e = sum(float(e.get('Valor', 0)) for e in st.session_state.efectivo)
        total_desglose = sub_t + sub_c + sub_g + sub_e
        venta_total = float(st.session_state.get('venta_total_dia', 0.0))
        diferencia = venta_total - total_desglose

        v1, v2, v3 = st.columns(3)
        v1.metric("💰 Venta Total (Sistema)", format_currency(venta_total))
        v2.metric("📊 Suma del Desglose", format_currency(total_desglose))
        delta_color = "inverse" if diferencia != 0 else "off"
        v3.metric("Diferencia", format_currency(diferencia), delta=format_currency(diferencia), delta_color=delta_color)

        if st.button("💾 Guardar o Actualizar Cuadre", type="primary", use_container_width=True):
            tienda = st.session_state.get("tienda_seleccionada")
            if not tienda:
                st.warning("🛑 Por favor, seleccione una tienda antes de guardar.")
                return
            if venta_total <= 0:
                st.warning("⚠️ La Venta Total del día debe ser mayor a cero.")
                return

            fecha_str = st.session_state.fecha_seleccionada.strftime("%d/%m/%Y")
            id_registro = f"{tienda}-{fecha_str}"

            try:
                cell = registros_ws.find(id_registro, in_column=1)
                
                if cell:
                    consecutivo_asignado_tienda = registros_ws.cell(cell.row, 2).value
                    consecutivo_global_doc = registros_ws.cell(cell.row, 15).value
                else:
                    consecutivo_asignado_tienda = get_next_consecutive(consecutivos_ws, tienda)
                    consecutivo_global_doc = get_next_global_consecutive(global_consecutivo_ws)
                    
                    if consecutivo_asignado_tienda is None or consecutivo_global_doc is None:
                        st.error("No se pudo generar uno de los consecutivos. No se guardará el registro.")
                        return
                    
                    update_consecutive(consecutivos_ws, tienda, consecutivo_asignado_tienda)
                    update_global_consecutive(global_consecutivo_ws, consecutivo_global_doc)

                fila_datos = [
                    id_registro, consecutivo_asignado_tienda, tienda, fecha_str,
                    st.session_state.factura_inicial, st.session_state.factura_final, venta_total,
                    json.dumps(st.session_state.tarjetas), json.dumps(st.session_state.consignaciones),
                    json.dumps(st.session_state.gastos), json.dumps(st.session_state.efectivo),
                    diferencia, datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "", 
                    consecutivo_global_doc
                ]

                if cell:
                    registros_ws.update(f'A{cell.row}', [fila_datos])
                    st.success(f"✅ Cuadre para {tienda} el {fecha_str} fue **actualizado**!")
                else:
                    registros_ws.append_row(fila_datos)
                    st.success(f"✅ Cuadre para {tienda} el {fecha_str} fue **guardado** con el consecutivo de referencia **{consecutivo_asignado_tienda}** y de documento **{consecutivo_global_doc}**!")
            except Exception as e:
                st.error(f"Error al guardar los datos en Google Sheets: {e}")

# --- 6. RENDERIZADO DE PÁGINAS PRINCIPALES ---
def render_form_page(worksheets, config):
    """Renderiza la página del formulario principal."""
    registros_ws, _, _, _ = worksheets
    tiendas, bancos, terceros = config
    
    st.header("1. Selección de Registro", anchor=False, divider="rainbow")
    c1,c2,c3,c4 = st.columns([2,2,1,1])
    c1.selectbox("Tienda", options=tiendas, key="tienda_seleccionada", on_change=clear_form_state, placeholder="Seleccione una tienda...")
    c2.date_input("Fecha", key="fecha_seleccionada", on_change=clear_form_state, format="DD/MM/YYYY")
    with c3:
        st.write(" ")
        st.button("🔍 Cargar Cuadre", on_click=load_cuadre_data, args=[registros_ws], use_container_width=True)
    with c4:
        st.write(" ")
        st.button("✨ Iniciar Nuevo", on_click=clear_form_state, use_container_width=True)

    st.divider()
    st.header("2. Formulario de Cuadre", anchor=False, divider="rainbow")
    
    with st.container(border=True):
        st.subheader("📋 Información General")
        c1,c2,c3=st.columns(3)
        st.session_state.factura_inicial=c1.text_input("Factura Inicial", value=st.session_state.get('factura_inicial', ""))
        st.session_state.factura_final=c2.text_input("Factura Final", value=st.session_state.get('factura_final', ""))
        st.session_state.venta_total_dia=c3.number_input("💰 Venta Total (Sistema)",min_value=0.0,step=1000.0,value=float(st.session_state.get('venta_total_dia', 0.0)),format="%.0f")

    with st.container(border=True):
        st.subheader("🧾 Desglose de Pagos")
        display_tarjetas_section()
        display_consignaciones_section(bancos)
        display_gastos_section(terceros)
        display_efectivo_section(terceros)

    display_summary_and_save(worksheets)

def render_reports_page(registros_ws, config_ws, tiendas_list):
    """Renderiza la página de generación de reportes."""
    st.header("Generación de Archivos y Reportes", divider="rainbow")
    st.markdown("Seleccione una tienda y un rango de fechas para generar los archivos para el sistema contable y los reportes de soporte.")

    today = datetime.now().date()
    col1, col2, col3 = st.columns(3)
    tienda_options = ["Todas las Tiendas"] + tiendas_list
    selected_store = col1.selectbox("Tienda", options=tienda_options)
    start_date = col2.date_input("Fecha de Inicio", today.replace(day=1))
    end_date = col3.date_input("Fecha de Fin", today)

    if start_date > end_date:
        st.error("Error: La fecha de inicio no puede ser posterior a la fecha de fin.")
        return

    st.divider()
    
    b1, b2, b3 = st.columns(3)

    with b1:
        if st.button("📄 Generar Archivo TXT", use_container_width=True, type="primary"):
            with st.spinner('Generando TXT...'):
                txt_content = generate_txt_file(registros_ws, config_ws, start_date, end_date, selected_store)
                if txt_content:
                    st.download_button(
                        label="📥 Descargar Archivo .txt",
                        data=txt_content.encode('utf-8'),
                        file_name=f"contabilidad_{selected_store.replace(' ','_')}_{start_date.strftime('%Y%m%d')}_a_{end_date.strftime('%Y%m%d')}.txt",
                        mime="text/plain",
                        use_container_width=True
                    )
                    st.success("Archivo TXT generado.")
    
    with b2:
        if st.button("📊 Generar Reporte Excel", use_container_width=True, type="primary"):
            with st.spinner('Creando un Excel impecable...'):
                excel_data = generate_excel_report(registros_ws, start_date, end_date, selected_store)
                if excel_data:
                    st.download_button(
                        label="📥 Descargar Reporte .xlsx",
                        data=excel_data,
                        file_name=f"Reporte_Cuadre_{selected_store.replace(' ','_')}_{start_date.strftime('%Y%m%d')}_a_{end_date.strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.success("Reporte Excel generado.")

    with b3:
        with st.form("email_form"):
            recipient_email = st.text_input("Email del Gerente", placeholder="ejemplo@dominio.com")
            submitted = st.form_submit_button("📧 Enviar Resumen Gerencial", use_container_width=True)
            
            if submitted:
                if recipient_email and "@" in recipient_email:
                    send_summary_email(registros_ws, start_date, end_date, selected_store, recipient_email)
                else:
                    st.warning("Por favor, ingrese una dirección de correo válida.")


# --- 7. FLUJO PRINCIPAL DE LA APLICACIÓN ---
def main():
    """Función principal que ejecuta la aplicación Streamlit."""
    st.title("CUADRE DIARIO DE CAJA")

    worksheets = connect_to_gsheet()
    
    if all(worksheets):
        registros_ws, config_ws, _, _ = worksheets
        with st.sidebar:
            st.header("Navegación")
            page_selection = st.radio(
                "Seleccione una página",
                ["📝 Formulario de Cuadre", "📈 Reportes"],
                key="page_radio",
                label_visibility="collapsed"
            )
            
            if page_selection == "📝 Formulario de Cuadre":
                st.session_state.page = "Formulario"
            else:
                st.session_state.page = "Reportes"
        
        config = get_app_config(config_ws)
        tiendas, _, _ = config

        if not tiendas and st.session_state.page == "Formulario":
            st.error("🚨 No se encontraron tiendas en la hoja de 'Configuracion'.")
            st.warning("Agregue al menos una tienda (Tipo Movimiento = TIENDA) para continuar.")
            return

        if st.session_state.page == "Formulario":
            render_form_page(worksheets, config)
        elif st.session_state.page == "Reportes":
            render_reports_page(registros_ws, config_ws, tiendas)
    else:
        st.info("⏳ Esperando conexión con Google Sheets...")

# --- BLOQUE DE EJECUCIÓN PRINCIPAL ---
if __name__ == "__main__":
    initialize_session_state()

    if check_password():
        main()
