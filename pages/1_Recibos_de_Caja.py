# -*- coding: utf-8 -*-

# --- IMPORTACIÓN DE LIBRERÍAS NECESARIAS ---
import streamlit as st
import pandas as pd
from io import BytesIO
from oauth2client.service_account import ServiceAccountCredentials
import gspread
from datetime import datetime, timedelta
from itertools import groupby
from operator import itemgetter
import time

# Importaciones para la generación y estilo del Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from app_shared import (
    current_authorized_series,
    filter_series_for_access,
    get_receipt_series_options,
    initialize_access_state,
    is_store_profile_active,
    render_sidebar,
    require_access,
    reset_session_state,
)

# --- CONFIGURACIÓN DE LA PÁGINA DE STREAMLIT ---
# Configura la página para que use un layout ancho y tenga un título.
st.set_page_config(layout="wide", page_title="Recibos de Caja")
initialize_access_state()
require_access(
    "store",
    "Recibos de caja",
    "Este modulo requiere perfil de tienda o acceso administrativo.",
)
render_sidebar("Recibos de caja")

# --- CREACIÓN DE PESTAÑAS (TABS) ---
# Aquí dividimos la aplicación en dos secciones grandes
tab1, tab2 = st.tabs(["📂 Procesamiento Recibos de Caja", "🏦 Identificación Bancaria"])

# ==============================================================================
# ========================= FUNCIONES GLOBALES / UTILS =========================
# ==============================================================================
# Estas funciones se usan en ambas pestañas o son de configuración general

# --- CONEXIÓN SEGURA A GOOGLE SHEETS ---
# Usa el cache de Streamlit para evitar reconectarse a Google Sheets en cada recarga de la página.
@st.cache_resource(ttl=600)
def connect_to_gsheet():
    """
    Establece una conexión con Google Sheets usando las credenciales de Streamlit.
    Devuelve los objetos de las hojas de cálculo necesarias.
    """
    try:
        # Cargar credenciales de los secrets de Streamlit.
        credenciales = st.secrets.get("google_credentials")
        if not credenciales:
             # Si no hay credenciales, retornamos None para evitar error si solo se usa la pestaña 2 sin internet
             return None, None, None, None

        creds_json = dict(credenciales)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_json, scope)
        client = gspread.authorize(creds)
        
        spreadsheet_name = "Planillas_Ferreinox"
        sheet = client.open(spreadsheet_name)
        
        # Accede a cada una de las hojas necesarias por su nombre.
        config_ws = sheet.worksheet("Configuracion")
        registros_recibos_ws = sheet.worksheet("RegistrosRecibos")
        consecutivos_ws = sheet.worksheet("Consecutivos")
        global_consecutivo_ws = sheet.worksheet("GlobalConsecutivo")
        
        return config_ws, registros_recibos_ws, consecutivos_ws, global_consecutivo_ws
        
    except gspread.exceptions.SpreadsheetNotFound:
        st.error(f"Error fatal: No se encontró el archivo de Google Sheets llamado '{spreadsheet_name}'. Revisa el nombre y los permisos.")
        return None, None, None, None
    except gspread.exceptions.WorksheetNotFound as e:
        st.error(f"Error fatal: No se encontró una de las hojas de trabajo requeridas. Detalle: {e}")
        st.warning("Asegúrate de que existan las hojas 'Configuracion', 'RegistrosRecibos', 'Consecutivos' y 'GlobalConsecutivo'.")
        return None, None, None, None
    except Exception as e:
        st.error(f"Error fatal al conectar con Google Sheets: {e}")
        st.warning("Verifica las credenciales en los secrets de Streamlit y los permisos de la cuenta de servicio.")
        return None, None, None, None

def get_app_config(config_ws):
    """
    Carga la configuración de bancos, terceros y destinos de tarjeta desde la hoja 'Configuracion'.
    """
    if config_ws is None:
        return [], [], {}, []
    try:
        config_data = config_ws.get_all_records()
        bancos = sorted(list(set(str(d['Detalle']).strip() for d in config_data if d.get('Tipo Movimiento') == 'BANCO' and d.get('Detalle'))))
        terceros = sorted(list(set(str(d['Detalle']).strip() for d in config_data if d.get('Tipo Movimiento') == 'TERCERO' and d.get('Detalle'))))
        
        tarjetas = sorted(list(set(str(d['Detalle']).strip() for d in config_data if d.get('Tipo Movimiento') == 'TARJETA' and d.get('Detalle'))))

        # Mapea los detalles a su información contable (cuenta, NIT, nombre).
        account_mappings = {}
        for d in config_data:
            detalle = str(d.get('Detalle', '')).strip()
            if detalle and (d.get('Tipo Movimiento') in ['BANCO', 'TERCERO', 'TARJETA']):
                account_mappings[detalle] = {
                    'cuenta': str(d.get('Cuenta Contable', '')).strip(),
                    'nit': str(d.get('NIT', '')).strip(),
                    'nombre': str(d.get('Nombre Tercero', '')).strip(),
                }
        return bancos, terceros, account_mappings, tarjetas
    except Exception as e:
        st.error(f"Error al cargar la configuración de bancos y terceros: {e}")
        return [], [], {}, []

# --- LÓGICA DE PROCESAMIENTO Y GENERACIÓN DE ARCHIVOS (TAB 1) ---
def generate_txt_content(df, account_mappings, tarjetas_destinos):
    """
    Genera el contenido del archivo TXT para el ERP.
    Agrupa por 'Consecutivo Global' para manejar cada lote diario de forma independiente.
    """
    txt_lines = []
    cuenta_recibo_caja = "11050501"
    tipo_documento = "12"

    if df.empty:
        return ""

    # Agrupa por el consecutivo global para procesar cada lote (diario) por separado.
    # Se usa sort=False para evitar reordenar si el DataFrame ya viene ordenado lógicamente.
    for global_consecutive, group_df in df.groupby('Consecutivo Global', sort=False):
        # Extrae datos comunes del lote.
        series_consecutive = group_df['Consecutivo Serie'].iloc[0]
        series = group_df['Serie'].iloc[0]
        series_numeric = ''.join(filter(str.isdigit, str(series)))

        # --- 1. PROCESAR REGISTROS INDIVIDUALES (DÉBITOS) ---
        df_individual = group_df[group_df['Agrupación'] == 1].copy()
        if not df_individual.empty:
            # Agrupa por el Recibo N° para sumar los valores que le corresponden (si hay múltiples detalles de factura por recibo)
            individual_grouped = df_individual.groupby('Recibo N°').agg(
                Valor_Total=('Valor Efectivo', 'sum'),
                Fecha=('Fecha', 'first'),
                Cliente=('Cliente', 'first'),
                Destino=('Destino', 'first')
            ).reset_index()

            for _, row in individual_grouped.iterrows():
                # Asegura que la fecha esté en formato DD/MM/AAAA
                try:
                    fecha = pd.to_datetime(row['Fecha'], dayfirst=True).strftime('%d/%m/%Y')
                except:
                    fecha = str(row['Fecha']) # En caso de error, usa la cadena original.
                    
                num_recibo = str(int(row['Recibo N°']))
                valor = float(row['Valor_Total'])
                destino = str(row['Destino'])
                
                serie_final_txt = str(series_numeric)
                # MODIFICACIÓN: Añadir 'Epayco' a la condición para el prefijo 'T'
                if destino in tarjetas_destinos or destino == 'Epayco':
                    serie_final_txt = "T" + serie_final_txt # Prefijo 'T' para tarjetas y Epayco.

                if destino in account_mappings:
                    destino_info = account_mappings[destino]
                    cuenta_destino = destino_info['cuenta']
                    nit_tercero = destino_info['nit']
                    nombre_tercero = destino_info['nombre']

                    linea_debito = "|".join([
                        fecha, str(global_consecutive), cuenta_destino, tipo_documento,
                        f"Recibo de Caja {num_recibo} - {row['Cliente']}",
                        serie_final_txt,
                        str(series_consecutive),
                        str(round(valor, 2)), "0", "0", nit_tercero, nombre_tercero, "0" # Valor en Débito
                    ])
                    txt_lines.append(linea_debito)

        # --- 2. PROCESAR REGISTROS AGRUPADOS (DÉBITOS) ---
        df_agrupado = group_df[group_df['Agrupación'] > 1]
        if not df_agrupado.empty:
            # Agrupa por Agrupación y Destino para generar una sola línea contable por grupo.
            grouped = df_agrupado.groupby(['Agrupación', 'Destino']).agg(
                Valor_Total=('Valor Efectivo', 'sum'),
                Fecha_Primera=('Fecha', 'first'),
                # Lista todos los recibos incluidos en este grupo, separados por coma.
                Recibos_Incluidos=('Recibo N°', lambda x: ','.join(sorted(list(set(x.astype(str).str.split('.').str[0])))))
            ).reset_index()

            for _, group_row in grouped.iterrows():
                destino = group_row['Destino']
                valor_total = group_row['Valor_Total']
                # Asegura que la fecha esté en formato DD/MM/AAAA
                try:
                    fecha = pd.to_datetime(group_row['Fecha_Primera'], dayfirst=True).strftime('%d/%m/%Y')
                except:
                    fecha = str(group_row['Fecha_Primera'])
                    
                recibos = group_row['Recibos_Incluidos']

                serie_final_txt = str(series_numeric)
                # MODIFICACIÓN: Añadir 'Epayco' a la condición para el prefijo 'T'
                if destino in tarjetas_destinos or destino == 'Epayco':
                    serie_final_txt = "T" + serie_final_txt

                if destino in account_mappings:
                    destino_info = account_mappings[destino]
                    cuenta_destino = destino_info['cuenta']
                    nit_tercero = destino_info['nit']
                    nombre_tercero = destino_info['nombre']
                    descripcion_grupo = f"Consolidado Recibos {recibos}"

                    linea_debito = "|".join([
                        fecha, str(global_consecutive), cuenta_destino, tipo_documento,
                        descripcion_grupo,
                        serie_final_txt,
                        str(series_consecutive),
                        str(round(valor_total, 2)), "0", "0", nit_tercero, nombre_tercero, "0" # Valor en Débito
                    ])
                    txt_lines.append(linea_debito)

        # --- 3. GENERAR LÍNEA DE CRÉDITO PARA EL LOTE DIARIO ---
        if not group_df.empty:
            total_dia = group_df['Valor Efectivo'].sum()
            
            try:
                fecha_cierre = pd.to_datetime(group_df['Fecha'].iloc[0], dayfirst=True).strftime('%d/%m/%Y')
            except:
                fecha_cierre = str(group_df['Fecha'].iloc[0])

            comentario_credito = f"Cierre Contable Fecha {fecha_cierre}"

            linea_credito_por_fecha = "|".join([
                fecha_cierre, str(global_consecutive), cuenta_recibo_caja, tipo_documento,
                comentario_credito,
                str(series_numeric), 
                str(series_consecutive),
                "0", str(round(total_dia, 2)), "0", "0", "0", "0" # Valor en Crédito
            ])
            txt_lines.append(linea_credito_por_fecha)

    return "\n".join(txt_lines)

# --- FUNCIÓN PARA GENERAR REPORTE EXCEL PROFESIONAL (CORREGIDA) ---
def generate_excel_report(df):
    """
    Genera un archivo Excel profesional y estilizado.
    Ordena primero por Fecha para manejar reportes de múltiples días.
    """
    output = BytesIO()
    
    # 1. Asegurar tipos de datos y crear columna temporal para ordenar.
    df['Recibo N°'] = pd.to_numeric(df['Recibo N°'], errors='coerce')
    df['Agrupación'] = pd.to_numeric(df['Agrupación'], errors='coerce')
    df['Valor Efectivo'] = pd.to_numeric(df['Valor Efectivo'], errors='coerce')
    df['Fecha_dt'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce')
    df.dropna(subset=['Recibo N°', 'Agrupación', 'Fecha_dt', 'Valor Efectivo'], inplace=True)
    
    # 2. Ordenar los datos usando la columna temporal.
    df.sort_values(by=['Fecha_dt', 'Agrupación', 'Recibo N°'], inplace=True)
    
    # 3. Formatear la columna de fecha a string para la visualización final.
    df['Fecha'] = df['Fecha_dt'].dt.strftime('%d/%m/%Y')

    # 4. Definir el orden final de columnas y seleccionar solo las necesarias.
    preferred_order = ['Fecha', 'Recibo N°', 'Serie-Número', 'Cliente', 'Valor Efectivo', 'Agrupación', 'Destino']
    
    # Construir la lista final de columnas, manteniendo el orden preferido al inicio.
    final_columns = preferred_order + [col for col in df.columns if col not in preferred_order and col != 'Fecha_dt']
    
    # Asegurarse de que el DataFrame solo contenga las columnas finales.
    df = df[final_columns]
    excel_columns = final_columns # Usar esta lista para el resto de la función.

    # Separar datos en individuales y grupos de consignación.
    df_individual = df[df['Agrupación'] == 1].copy()
    df_grouped = df[df['Agrupación'] > 1].copy()

    report_data = []

    # Procesar recibos individuales con subtotales.
    if not df_individual.empty:
        for recibo_num, group in df_individual.groupby('Recibo N°', sort=False):
            for _, row in group.iterrows():
                report_data.append(row[excel_columns].tolist())
            
            subtotal = group['Valor Efectivo'].sum()
            subtotal_row = [''] * len(excel_columns)
            cliente_col_idx = excel_columns.index('Cliente')
            valor_col_idx = excel_columns.index('Valor Efectivo')
            subtotal_row[cliente_col_idx] = f'Subtotal Recibo N° {int(recibo_num)}'
            subtotal_row[valor_col_idx] = subtotal
            report_data.append(subtotal_row)

    # Procesar consignaciones agrupadas con subtotales.
    if not df_grouped.empty:
        for agrupacion_id, group in df_grouped.groupby('Agrupación', sort=False):
            for _, row in group.iterrows():
                report_data.append(row[excel_columns].tolist())
            
            subtotal = group['Valor Efectivo'].sum()
            subtotal_row = [''] * len(excel_columns)
            cliente_col_idx = excel_columns.index('Cliente')
            valor_col_idx = excel_columns.index('Valor Efectivo')
            subtotal_row[cliente_col_idx] = f'Subtotal Consignación Grupo {int(agrupacion_id)}'
            subtotal_row[valor_col_idx] = subtotal
            report_data.append(subtotal_row)
    
    report_df = pd.DataFrame(report_data, columns=excel_columns) if report_data else pd.DataFrame(columns=excel_columns)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        report_df.to_excel(writer, index=False, sheet_name='Recibos de Caja')
        workbook = writer.book
        worksheet = writer.sheets['Recibos de Caja']

        # --- Definición de Estilos ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        subtotal_font = Font(bold=True)
        subtotal_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        currency_format = '$ #,##0.00'

        # Aplicar estilo al encabezado.
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Aplicar estilo a las filas de datos y subtotales.
        valor_col_letter = get_column_letter(excel_columns.index('Valor Efectivo') + 1)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            is_subtotal_row = False
            cliente_cell_value = str(row[excel_columns.index('Cliente')].value) if row[excel_columns.index('Cliente')].value is not None else ''
            if cliente_cell_value.startswith('Subtotal'):
                is_subtotal_row = True
            
            for cell in row:
                cell.border = thin_border
                if is_subtotal_row:
                    cell.font = subtotal_font
                    cell.fill = subtotal_fill
            
            # Formatear y alinear celdas numéricas y de texto.
            valor_cell = worksheet[f'{valor_col_letter}{row_idx}']
            if isinstance(valor_cell.value, (int, float)):
                valor_cell.number_format = currency_format

            for col_name, align in [('Recibo N°', 'center'), ('Valor Efectivo', 'right'), ('Agrupación', 'center'), ('Serie-Número', 'center')]:
                if col_name in excel_columns:
                    col_idx = excel_columns.index(col_name) + 1
                    worksheet.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal=align)
        
        # --- Añadir Fila de Total General ---
        grand_total = df['Valor Efectivo'].sum()
        total_row_idx = worksheet.max_row + 1
        
        # Buscamos los índices de columna para 'Cliente' y 'Valor Efectivo'
        cliente_col_idx = excel_columns.index('Cliente') + 1
        valor_col_idx = excel_columns.index('Valor Efectivo') + 1
        
        worksheet.cell(row=total_row_idx, column=cliente_col_idx, value='TOTAL GENERAL')
        total_valor_cell = worksheet.cell(row=total_row_idx, column=valor_col_idx, value=grand_total)
        
        for cell in worksheet[total_row_idx]:
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
        
        total_valor_cell.number_format = currency_format
        total_valor_cell.alignment = Alignment(horizontal='right')

        # --- Ajustar el ancho de las columnas ---
        for col_idx, column_cells in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    # Usar la longitud de la celda de cabecera como mínimo
                    current_length = len(str(cell.value))
                    if cell.row == 1:
                        current_length = max(len(str(column_cells[0].value)), current_length)
                        
                    if current_length > max_length:
                        max_length = current_length
                except:
                    pass
            # Añadir un margen de 2 para mejor visualización, y asegurar un mínimo de 10.
            adjusted_width = max(12, (max_length + 2)) 
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return output.getvalue()


# --- FUNCIONES PARA MANEJAR CONSECUTIVOS ---
def get_next_series_consecutive(consecutivos_ws, series_name):
    """Obtiene el siguiente número consecutivo para una serie específica."""
    try:
        # Busca la etiqueta en la hoja
        cell = consecutivos_ws.find(f'Ultimo_Consecutivo_{series_name}')
        if cell:
            # El valor está en la columna siguiente
            return int(consecutivos_ws.cell(cell.row, cell.col + 1).value) + 1
        st.error(f"No se encontró la etiqueta para la serie '{series_name}'. Revisa la hoja 'Consecutivos'.")
        return None
    except Exception as e:
        st.error(f"Error obteniendo el consecutivo para la serie {series_name}: {e}")
        return None

def update_series_consecutive(consecutivos_ws, series_name, new_consecutive):
    """Actualiza el último número consecutivo utilizado para una serie."""
    try:
        cell = consecutivos_ws.find(f'Ultimo_Consecutivo_{series_name}')
        if cell:
            consecutivos_ws.update_cell(cell.row, cell.col + 1, new_consecutive)
    except Exception as e:
        st.error(f"Error actualizando el consecutivo para la serie {series_name}: {e}")

def get_next_global_consecutive(global_consecutivo_ws):
    """Obtiene el siguiente número consecutivo global."""
    try:
        # Asume que el consecutivo global está en B1
        return int(global_consecutivo_ws.acell('B1').value) + 1
    except Exception as e:
        st.error(f"Error obteniendo el consecutivo global: {e}")
        return None

def update_global_consecutive(global_consecutivo_ws, new_consecutive):
    """Actualiza el último número consecutivo global."""
    try:
        global_consecutivo_ws.update_acell('B1', new_consecutive)
    except Exception as e:
        st.error(f"Error actualizando el consecutivo global: {e}")

# --- FUNCIÓN PARA BORRAR REGISTROS (MODIFICADA para borrar por lista de CONSECUTIVOS GLOBALES) ---
def delete_existing_records(ws, global_consecutives_to_delete):
    """
    Encuentra y borra todas las filas que coincidan con una lista de consecutivos globales
    utilizando una solicitud por lotes (batch) para evitar errores de cuota.
    """
    try:
        if not global_consecutives_to_delete:
            st.warning("No hay consecutivos globales especificados para borrar.")
            return
        
        consecutives_str = [str(c) for c in global_consecutives_to_delete]
        st.info(f"Buscando registros antiguos con consecutivos globales: {', '.join(consecutives_str)} para eliminarlos...")
        
        # Obtener todos los valores.
        all_records = ws.get_all_values() 
        if len(all_records) <= 1:
            st.warning("No hay registros en la hoja para buscar. Se procederá a guardar como si fueran nuevos.")
            return

        headers = all_records[0]
        df_records = pd.DataFrame(all_records[1:], columns=headers)
        
        if 'Consecutivo Global' not in df_records.columns:
            st.error("La hoja 'RegistrosRecibos' no tiene la columna 'Consecutivo Global'. No se puede actualizar.")
            return

        df_records['Consecutivo Global'] = df_records['Consecutivo Global'].astype(str)
        
        # Encontrar los índices de las filas a borrar
        rows_to_delete_indices = df_records[df_records['Consecutivo Global'].isin(consecutives_str)].index.tolist()
        
        if not rows_to_delete_indices:
            st.warning("No se encontraron registros antiguos que coincidieran. Se procederá a guardar como si fueran nuevos.")
            return

        # Convertir los índices de pandas a los índices de fila de gspread (base 1, +1 por cabecera).
        gspread_rows_to_delete = sorted([i + 2 for i in rows_to_delete_indices])

        # Agrupa índices consecutivos para minimizar las solicitudes de borrado.
        requests = []
        for _, g in groupby(enumerate(gspread_rows_to_delete), lambda i_x: i_x[0] - i_x[1]):
            group = list(map(itemgetter(1), g))
            requests.append({
                "deleteDimension": {
                    "range": {
                        "sheetId": ws.id,
                        "dimension": "ROWS",
                        "startIndex": group[0] - 1, # El API es base 0
                        "endIndex": group[-1]
                    }
                }
            })
        
        if requests:
            # Las solicitudes de borrado deben ir de abajo hacia arriba para no alterar los índices de las filas superiores.
            requests.reverse()
            ws.spreadsheet.batch_update({"requests": requests})
            st.success(f"Se eliminaron {len(gspread_rows_to_delete)} registros antiguos en una sola operación por lotes.")

    except Exception as e:
        st.error(f"Error crítico al intentar borrar registros antiguos: {e}")
        st.stop()


# ==============================================================================
# ========================= PESTAÑA 1: TU CÓDIGO ORIGINAL ======================
# ==============================================================================
with tab1:
    # --- LÓGICA PRINCIPAL DE LA PÁGINA ---
    # Títulos y descripción original
    st.title("🧾 Procesamiento de Recibos de Caja v5.8 (Corregido)")
    st.markdown("""
    Esta herramienta ahora permite tres flujos de trabajo:
    1.  **Descargar reportes antiguos**: Busca y descarga un **reporte consolidado** con todos los grupos procesados en un rango de fechas y serie.
    2.  **Cargar un nuevo archivo de Excel**: Procesa un nuevo grupo de recibos, asignando **consecutivos por día** si el archivo abarca varias fechas, y lo guarda generando un reporte detallado.
    3.  **Buscar y editar un grupo existente**: Carga un grupo completo (incluso con fechas diferentes), permite **editar el valor del recibo** y volver a guardarlo, respetando la serie.
    """)

    config_ws, registros_recibos_ws, consecutivos_ws, global_consecutivo_ws = connect_to_gsheet()

    if any(ws is None for ws in [config_ws, registros_recibos_ws, consecutivos_ws, global_consecutivo_ws]):
        st.error("La aplicación no puede continuar debido a un error de conexión con Google Sheets.")
    else:
        bancos, terceros, account_mappings, tarjetas_destinos = get_app_config(config_ws)
        opciones_destino = ["-- Seleccionar --"] + bancos + terceros + tarjetas_destinos
        opciones_agrupacion = list(range(1, 11))
        series_disponibles = filter_series_for_access(get_receipt_series_options())
        store_series_locked = is_store_profile_active() and bool(current_authorized_series())

        if not series_disponibles:
            st.error("El perfil actual no tiene una serie autorizada configurada.")
            st.stop()
        
        # Inicializa el estado de la sesión si no existe.
        if 'mode' not in st.session_state:
            st.session_state.mode = 'new'
            st.session_state.editing_info = {}
            st.session_state.found_groups = []
            st.session_state.df_full_detail = pd.DataFrame() # DataFrame completo del detalle de facturas
            st.session_state.df_for_display = pd.DataFrame() # DataFrame resumido para la tabla de edición

        # --- SECCIÓN DE DESCARGA DE REPORTES ANTERIORES ---
        st.divider()
        with st.expander("📥 Descargar Reportes Anteriores", expanded=False):
            st.info("Busca todos los grupos dentro de un rango de fechas y serie para generar y descargar un **reporte consolidado**.")
            
            dl_col1, dl_col2, dl_col3 = st.columns(3)
            with dl_col1:
                start_date = st.date_input("Fecha de inicio:", datetime.now().date(), key="dl_start_date")
            with dl_col2:
                end_date = st.date_input("Fecha de fin:", datetime.now().date(), key="dl_end_date")
            with dl_col3:
                download_serie = st.selectbox(
                    "Serie a buscar:",
                    options=series_disponibles,
                    key="dl_serie",
                    disabled=store_series_locked,
                )

            if store_series_locked:
                st.caption(f"Perfil restringido a la serie: {', '.join(series_disponibles)}")
            
            if st.button("Buscar y Preparar Reporte Consolidado", use_container_width=True):
                if end_date < start_date:
                    st.error("Error: La fecha de fin no puede ser anterior a la fecha de inicio.")
                else:
                    try:
                        with st.spinner("Buscando registros en Google Sheets..."):
                            all_values = registros_recibos_ws.get_all_values()
                            if len(all_values) > 1:
                                headers = all_values[0]
                                all_records_df = pd.DataFrame(all_values[1:], columns=headers)
                                
                                # Limpieza de datos
                                all_records_df = all_records_df.drop(columns=[''], errors='ignore')
                                all_records_df['Fecha_dt'] = pd.to_datetime(all_records_df['Fecha'], format='%d/%m/%Y', errors='coerce')
                                
                                # Filtrar por rango de fechas y serie. Se usa <= en el final para ser inclusivo.
                                filtered_df = all_records_df[
                                    (all_records_df['Fecha_dt'].dt.date >= start_date) &
                                    (all_records_df['Fecha_dt'].dt.date <= end_date) &
                                    (all_records_df['Serie'] == download_serie)
                                ].copy()

                                # Se asegura que TODAS las columnas numéricas clave se conviertan aquí.
                                if not filtered_df.empty:
                                    numeric_cols = ['Valor Efectivo', 'Agrupación', 'Recibo N°']
                                    for col in numeric_cols:
                                        filtered_df[col] = pd.to_numeric(filtered_df[col], errors='coerce')
                                    
                                    # Eliminar filas donde la conversión numérica o de fecha falló.
                                    filtered_df.dropna(subset=['Fecha_dt'] + numeric_cols, inplace=True)

                                if not filtered_df.empty:
                                    st.session_state.df_for_consolidated_download = filtered_df
                                    st.success(f"¡Búsqueda exitosa! Se encontraron {len(filtered_df['Consecutivo Global'].unique())} grupos. El reporte consolidado está listo para descargar.")
                                else:
                                    st.warning("No se encontraron grupos válidos para el rango de fechas y serie seleccionados.")
                                    if 'df_for_consolidated_download' in st.session_state:
                                        del st.session_state.df_for_consolidated_download
                            else:
                                st.warning("No hay registros guardados para buscar.")
                    except Exception as e:
                        st.error(f"Ocurrió un error al buscar los registros: {e}")

        if 'df_for_consolidated_download' in st.session_state and not st.session_state.df_for_consolidated_download.empty:
            df_for_download = st.session_state.df_for_consolidated_download.copy()
            
            # Crear columna 'Serie-Número' para el reporte.
            s_factura = df_for_download['Serie_Factura'].fillna('S/D').astype(str)
            n_factura = df_for_download['Numero_Factura'].fillna('S/D').astype(str)
            df_for_download['Serie-Número'] = s_factura + "-" + n_factura

            # Generar archivos consolidados.
            txt_content_dl = generate_txt_content(df_for_download.copy(), account_mappings, tarjetas_destinos)
            excel_file_dl = generate_excel_report(df_for_download.copy())
            
            dl_btn_col1, dl_btn_col2 = st.columns(2)
            with dl_btn_col1:
                st.download_button(
                    label="⬇️ Descargar Archivo TXT Consolidado",
                    data=txt_content_dl.encode('utf-8'),
                    file_name=f"recibos_consolidados_{download_serie}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.txt",
                    mime="text/plain", use_container_width=True
                )
            with dl_btn_col2:
                st.download_button(
                    label="📄 Descargar Reporte Excel Consolidado",
                    data=excel_file_dl,
                    file_name=f"Reporte_Recibos_Consolidado_{download_serie}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True
                )
        st.divider()

        # --- SECCIÓN PRINCIPAL DE PROCESAMIENTO ---
        st.header("Flujo de Trabajo: Procesar o Editar")
        st.subheader("1. Elige una opción")

        col_mode_1, col_mode_2, _ = st.columns([1,1,2])
        # Botones para cambiar entre modo 'nuevo' y 'editar'.
        if col_mode_1.button("🆕 Procesar Nuevo Archivo", use_container_width=True, type="primary" if st.session_state.mode == 'new' else "secondary"):
            # Limpia el estado de la sesión para empezar de cero.
            reset_session_state('mode', 'google_credentials')
            st.session_state.mode = 'new'
            st.session_state.df_full_detail = pd.DataFrame() # Reiniciar DataFrames
            st.session_state.df_for_display = pd.DataFrame()
            st.rerun()

        if col_mode_2.button("✏️ Editar Grupo Existente", use_container_width=True, type="primary" if st.session_state.mode == 'edit' else "secondary"):
            # Limpia el estado de la sesión para empezar de cero.
            reset_session_state('mode', 'google_credentials')
            st.session_state.mode = 'edit'
            st.session_state.df_full_detail = pd.DataFrame() # Reiniciar DataFrames
            st.session_state.df_for_display = pd.DataFrame()
            st.rerun()
            
        # --- MODO EDICIÓN: BUSCAR Y CARGAR GRUPO ---
        if st.session_state.mode == 'edit':
            st.subheader("2. Buscar y Cargar Registros para Edición")
            st.info("Busca **todos los recibos** dentro de un rango de fechas y una serie específica. Al guardar, se mantendrán los consecutivos originales.")
            
            with st.container(border=True):
                search_col1, search_col2, search_col3 = st.columns(3)
                with search_col1:
                    search_start_date = st.date_input("Fecha de inicio:", datetime.now().date(), key="edit_start_date")
                with search_col2:
                    search_end_date = st.date_input("Fecha de fin:", datetime.now().date(), key="edit_end_date")
                with search_col3:
                    search_serie = st.selectbox(
                        "Serie de los recibos:",
                        options=series_disponibles,
                        key="search_serie",
                        disabled=store_series_locked,
                    )

                if store_series_locked:
                    st.caption(f"Perfil restringido a la serie: {', '.join(series_disponibles)}")
                
                if st.button("Cargar Registros para Editar", use_container_width=True, type="primary"):
                    if search_end_date < search_start_date:
                        st.error("Error: La fecha de fin no puede ser anterior a la fecha de inicio.")
                    else:
                        try:
                            with st.spinner("Buscando, por favor espera..."):
                                all_values = registros_recibos_ws.get_all_values()
                                
                                if len(all_values) < 2:
                                    st.warning("No hay registros en la hoja para buscar.")
                                    st.session_state.df_full_detail = pd.DataFrame()
                                    st.session_state.df_for_display = pd.DataFrame()
                                else:
                                    headers = all_values[0]
                                    all_records_df = pd.DataFrame(all_values[1:], columns=headers)
                                    all_records_df = all_records_df.drop(columns=[''], errors='ignore')
                                    
                                    # Convertir fecha para poder comparar rangos
                                    all_records_df['Fecha_dt'] = pd.to_datetime(all_records_df['Fecha'], format='%d/%m/%Y', errors='coerce')
                                    all_records_df.dropna(subset=['Fecha_dt'], inplace=True)
                                    
                                    # MODIFICACIÓN: Filtrar también por la serie seleccionada.
                                    filtered_df = all_records_df[
                                        (all_records_df['Fecha_dt'].dt.date >= search_start_date) & 
                                        (all_records_df['Fecha_dt'].dt.date <= search_end_date) &
                                        (all_records_df['Serie'] == search_serie)
                                    ].copy()
                                    
                                    # Asegurar columnas numéricas para el detalle
                                    for col in ['Valor Efectivo', 'Agrupación', 'Consecutivo Global', 'Consecutivo Serie']:
                                        filtered_df[col] = pd.to_numeric(filtered_df[col], errors='coerce')

                                    if not filtered_df.empty:
                                        # Guardar el detalle completo (Múltiples líneas por recibo si es el caso)
                                        # Renombrar para consistencia con el flujo de 'nuevo archivo'
                                        filtered_df.rename(columns={
                                            'Serie_Factura': 'SERIE_FACTURA',
                                            'Numero_Factura': 'NUMERO_FACTURA'
                                            }, inplace=True)
                                        st.session_state.df_full_detail = filtered_df.copy()
                                        
                                        # Obtener la lista única de consecutivos globales a borrar/reemplazar.
                                        global_consecutives = st.session_state.df_full_detail['Consecutivo Global'].unique().tolist()
                                        series_consecutives = st.session_state.df_full_detail['Consecutivo Serie'].unique().tolist()
                                        
                                        # Crear el resumen para la tabla de edición (un resumen por Recibo N°).
                                        df_summary_edit = filtered_df.groupby('Recibo N°').agg(
                                            Fecha=('Fecha', 'first'),
                                            Cliente=('Cliente', 'first'),
                                            Valor_Efectivo_Total=('Valor Efectivo', 'sum'),
                                            Agrupación=('Agrupación', 'first'),
                                            Destino=('Destino', 'first')
                                        ).reset_index()
                                        df_summary_edit.rename(columns={'Valor_Efectivo_Total': 'Valor Efectivo'}, inplace=True)
                                        st.session_state.df_for_display = df_summary_edit[['Fecha', 'Recibo N°', 'Cliente', 'Valor Efectivo', 'Agrupación', 'Destino']]
                                        
                                        st.session_state.editing_info = {
                                            # Almacena la lista de todos los consecutivos globales a mantener.
                                            'global_consecutives': global_consecutives,
                                            'series_consecutives': series_consecutives,
                                            'serie': search_serie
                                        }
                                        st.success(f"Se cargaron {len(df_summary_edit)} recibos de caja ({len(global_consecutives)} grupos). Ahora puedes editarlos en la tabla de abajo.")
                                        st.rerun() # Dispara rerun para limpiar los elementos anteriores.
                                    else:
                                        st.session_state.df_full_detail = pd.DataFrame()
                                        st.session_state.df_for_display = pd.DataFrame()
                                        st.warning("No se encontraron registros para ese rango de fechas y serie.")
                        except Exception as e:
                            st.error(f"Error al buscar registros: {e}")
                            st.session_state.df_full_detail = pd.DataFrame()
                            st.session_state.df_for_display = pd.DataFrame()


        # --- MODO NUEVO: CARGAR ARCHIVO EXCEL ---
        elif st.session_state.mode == 'new':
            st.subheader("2. Cargar Nuevo Archivo")
            
            with st.container(border=True):
                st.markdown("##### A. Selecciona la Serie del Documento")
                serie_seleccionada = st.selectbox(
                    "Elige la serie que corresponde a los recibos de este archivo:",
                    options=series_disponibles,
                    index=0,
                    key="new_serie_select",
                    disabled=store_series_locked,
                )

                if store_series_locked:
                    st.caption(f"Perfil restringido a la serie: {', '.join(series_disponibles)}")
                
                st.markdown("##### B. Carga el Archivo de Excel")
                uploaded_file = st.file_uploader(
                    "📂 Sube tu archivo de Excel de recibos de caja",
                    type=['xlsx', 'xls']
                )

            if uploaded_file and ('df_for_display' not in st.session_state or st.session_state.get('uploaded_file_name') != uploaded_file.name):
                with st.spinner("Procesando archivo de Excel..."):
                    try:
                        # Lee el archivo, quitando la última fila que suele ser un total.
                        df = pd.read_excel(uploaded_file, header=0).iloc[:-1]
                        # Normaliza nombres de columna
                        df.columns = df.columns.astype(str).str.strip().str.upper().str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8')
                        
                        column_mapping = {
                            'NUMRECIBO': ['NUMRECIBO', 'RECIBO'], 'NOMBRECLIENTE': ['NOMBRECLIENTE', 'CLIENTE'],
                            'FECHA_RECIBO': ['FECHA_RECIBO', 'FECHA'], 'IMPORTE': ['IMPORTE', 'VALOR'],
                            'NUMERO_FACTURA': ['NUMERO'], 'SERIE_FACTURA': ['SERIE']
                        }
                        # Busca columnas existentes basadas en el mapeo
                        found_columns = {name: std_name for std_name, names in column_mapping.items() for name in names if name in df.columns}
                        df.rename(columns=found_columns, inplace=True)
                        
                        # Validar columnas y limpiar datos
                        df_cleaned = df.dropna(subset=['IMPORTE']).copy()
                        # Rellenar valores nulos de recibo, fecha y cliente con el valor anterior
                        for col in ['NUMRECIBO', 'FECHA_RECIBO', 'NOMBRECLIENTE']:
                            df_cleaned[col] = df_cleaned[col].ffill()
                        
                        # Limpieza del valor de importe (quitando $ y usando punto como decimal)
                        df_cleaned['IMPORTE_LIMPIO'] = df_cleaned['IMPORTE'].apply(
                            lambda x: pd.to_numeric(str(x).replace('$', '').strip().replace('.', '').replace(',', '.'), errors='coerce')
                        )
                        df_cleaned.dropna(subset=['IMPORTE_LIMPIO'], inplace=True)

                        df_full_detail = df_cleaned.rename(columns={
                            'FECHA_RECIBO': 'Fecha', 'NUMRECIBO': 'Recibo N°',
                            'NOMBRECLIENTE': 'Cliente', 'IMPORTE_LIMPIO': 'Valor Efectivo'
                        })
                        
                        # Asegura que la columna de fecha esté en el formato 'DD/MM/AAAA'
                        if pd.api.types.is_datetime64_any_dtype(df_full_detail['Fecha']):
                            df_full_detail['Fecha'] = pd.to_datetime(df_full_detail['Fecha']).dt.strftime('%d/%m/%Y')
                        else:
                            # Intenta convertir a string y luego a datetime para aplicar formato DD/MM/AAAA, si falla lo deja como estaba
                            try:
                                df_full_detail['Fecha'] = df_full_detail['Fecha'].astype(str)
                                df_full_detail['Fecha'] = pd.to_datetime(df_full_detail['Fecha'], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y').fillna(df_full_detail['Fecha'])
                            except:
                                pass
                        
                        st.session_state.df_full_detail = df_full_detail.copy()

                        # Crea el DataFrame resumido para la edición (un fila por recibo N°)
                        df_summary = df_full_detail.groupby('Recibo N°').agg(
                            Fecha=('Fecha', 'first'),
                            Cliente=('Cliente', 'first'),
                            Valor_Efectivo_Total=('Valor Efectivo', 'sum')
                        ).reset_index()
                        df_summary.rename(columns={'Valor_Efectivo_Total': 'Valor Efectivo'}, inplace=True)
                        df_summary['Agrupación'] = 1 # Valor predeterminado
                        df_summary['Destino'] = "-- Seleccionar --" # Valor predeterminado
                        
                        st.session_state.df_for_display = df_summary[['Fecha', 'Recibo N°', 'Cliente', 'Valor Efectivo', 'Agrupación', 'Destino']]
                        st.session_state.uploaded_file_name = uploaded_file.name
                        st.session_state.editing_info = {'serie': serie_seleccionada}
                        st.success("¡Archivo procesado! Ahora puedes asignar destinos y grupos.")
                        st.rerun()

                    except Exception as e:
                        st.error(f"Ocurrió un error al leer o procesar el archivo de Excel: {e}")
                        st.session_state.df_full_detail = pd.DataFrame()
                        st.session_state.df_for_display = pd.DataFrame()


        # --- TABLA DE EDICIÓN Y PROCESAMIENTO (COMÚN PARA AMBOS MODOS) ---
        if 'df_for_display' in st.session_state and not st.session_state.df_for_display.empty:
            st.divider()
            st.header("3. Asigna Agrupación y Destinos")
            
            st.metric(label="💰 Total Efectivo del Grupo", value=f"${st.session_state.df_for_display['Valor Efectivo'].sum():,.2f}")

            # Herramientas de asignación masiva.
            with st.expander("Herramientas de asignación masiva"):
                col1, col2 = st.columns(2)
                with col1:
                    destino_masivo = st.selectbox("Asignar destino a todos:", options=opciones_destino, key="masive_destino")
                    if st.button("Aplicar Destino", use_container_width=True) and destino_masivo != "-- Seleccionar --":
                        st.session_state.df_for_display['Destino'] = destino_masivo
                        st.rerun()
                with col2:
                    agrupacion_masiva = st.selectbox("Asignar grupo a todos:", options=opciones_agrupacion, key="masive_agrupacion")
                    if st.button("Aplicar Grupo", use_container_width=True):
                        st.session_state.df_for_display['Agrupación'] = agrupacion_masiva
                        st.rerun()
            
            st.info("Edita la agrupación, el destino y el valor para cada recibo. El detalle completo se usará para el reporte final.")
            
            # Tabla editable para que el usuario asigne grupos y destinos.
            edited_summary_df = st.data_editor(
                st.session_state.df_for_display,
                column_config={
                    "Agrupación": st.column_config.SelectboxColumn("Agrupación", help="Grupo 1 es individual. Grupos >1 se sumarán.", options=opciones_agrupacion, required=True),
                    "Destino": st.column_config.SelectboxColumn("Destino del Efectivo", help="Selecciona el banco o tercero.", options=opciones_destino, required=True),
                    # MODIFICACIÓN: Permitir la edición del valor efectivo.
                    "Valor Efectivo": st.column_config.NumberColumn("Valor Total Recibo", format="$ %.2f", required=True, disabled=False),
                    "Fecha": st.column_config.TextColumn("Fecha", disabled=True),
                    "Cliente": st.column_config.TextColumn("Cliente", disabled=True),
                    "Recibo N°": st.column_config.NumberColumn("Recibo N°", disabled=True),
                },
                hide_index=True, use_container_width=True,
                column_order=['Fecha', 'Recibo N°', 'Cliente', 'Valor Efectivo', 'Agrupación', 'Destino']
            )
            
            st.divider()
            st.header("4. Finalizar Proceso")
            
            if st.button("💾 Procesar y Guardar Cambios", type="primary", use_container_width=True):
                if any(d == "-- Seleccionar --" for d in edited_summary_df['Destino']):
                    st.warning("⚠️ Debes asignar un destino válido para TODOS los recibos antes de procesar.")
                else:
                    with st.spinner("Guardando datos y generando archivos..."):
                        try:
                            serie_seleccionada = st.session_state.editing_info['serie']
                            
                            # --- MODIFICACIÓN: Lógica para actualizar los valores de efectivo si fueron editados ---
                            # Compara el DataFrame editado con el original para encontrar cambios en 'Valor Efectivo'.
                            comparison_df = pd.merge(
                                st.session_state.df_for_display.rename(columns={'Valor Efectivo': 'Valor Efectivo_orig'}),
                                edited_summary_df.rename(columns={'Valor Efectivo': 'Valor Efectivo_edit'}),
                                on=['Recibo N°', 'Fecha', 'Cliente'] # Usar más claves para una unión robusta
                            )
                            changed_values = comparison_df[comparison_df['Valor Efectivo_orig'] != comparison_df['Valor Efectivo_edit']]

                            if not changed_values.empty:
                                st.info("Actualizando valores de efectivo editados en el detalle...")
                                df_full_detail_copy = st.session_state.df_full_detail.copy()
                                for _, row in changed_values.iterrows():
                                    recibo_num = row['Recibo N°']
                                    new_value = row['Valor Efectivo_edit']
                                    
                                    matching_indices = df_full_detail_copy[df_full_detail_copy['Recibo N°'] == recibo_num].index
                                    
                                    if not matching_indices.empty:
                                        # Pone el nuevo valor total en la primera fila de detalle del recibo.
                                        df_full_detail_copy.loc[matching_indices[0], 'Valor Efectivo'] = new_value
                                        # Si hay más filas de detalle para el mismo recibo, las pone en cero para que la suma total coincida.
                                        if len(matching_indices) > 1:
                                            df_full_detail_copy.loc[matching_indices[1:], 'Valor Efectivo'] = 0
                                    
                                # Actualiza el DataFrame de detalle en la sesión.
                                st.session_state.df_full_detail = df_full_detail_copy
                                st.success("Valores de efectivo actualizados en el detalle.")

                            # --- Lógica Común: Fusionar el resumen editado con el detalle completo ---
                            df_to_update = st.session_state.df_full_detail.drop(columns=['Agrupación', 'Destino', 'Serie'], errors='ignore')
                            
                            # Fusionar los nuevos campos de 'Agrupación' y 'Destino'
                            df_full_detail_merged = pd.merge(
                                df_to_update,
                                edited_summary_df[['Recibo N°', 'Agrupación', 'Destino']],
                                on='Recibo N°',
                                how='left'
                            )
                            df_full_detail_merged['Serie'] = serie_seleccionada


                            if st.session_state.mode == 'new':
                                st.info("Procesando como un NUEVO grupo con consecutivos diarios...")
                                
                                processed_daily_dfs = []
                                # Generar un Consecutivo Global y de Serie NUEVOS para cada fecha única.
                                for date_str in sorted(df_full_detail_merged['Fecha'].unique()):
                                    global_consecutive = get_next_global_consecutive(global_consecutivo_ws)
                                    series_consecutive = get_next_series_consecutive(consecutivos_ws, serie_seleccionada)

                                    if global_consecutive is None or series_consecutive is None:
                                        st.error(f"No se pudieron obtener los consecutivos para la fecha {date_str}. Revisa la configuración en Google Sheets.")
                                        st.stop()

                                    daily_df = df_full_detail_merged[df_full_detail_merged['Fecha'] == date_str].copy()
                                    daily_df['Consecutivo Global'] = global_consecutive
                                    daily_df['Consecutivo Serie'] = series_consecutive
                                    processed_daily_dfs.append(daily_df)
                                    
                                    # Actualizar consecutivos para el siguiente día.
                                    update_global_consecutive(global_consecutivo_ws, global_consecutive)
                                    update_series_consecutive(consecutivos_ws, serie_seleccionada, series_consecutive)
                                
                                final_df_to_process = pd.concat(processed_daily_dfs)

                            elif st.session_state.mode == 'edit':
                                st.info("Procesando como una EDICIÓN de grupo existente, manteniendo consecutivos...")
                                
                                # 1. BORRAR TODOS LOS REGISTROS CON LOS CONSECUTIVOS GLOBALES ENCONTRADOS
                                global_consecutives_to_delete = st.session_state.editing_info['global_consecutives']
                                delete_existing_records(registros_recibos_ws, global_consecutives_to_delete)

                                # 2. MANTENER LOS CONSECUTIVOS ORIGINALES
                                # Los campos 'Consecutivo Global' y 'Consecutivo Serie' ya están en df_full_detail_merged
                                # porque venían en los datos cargados desde Google Sheets.
                                final_df_to_process = df_full_detail_merged
                                # Nos aseguramos de que sean numéricos para el TXT/Excel
                                final_df_to_process['Consecutivo Global'] = pd.to_numeric(final_df_to_process['Consecutivo Global'], errors='coerce')
                                final_df_to_process['Consecutivo Serie'] = pd.to_numeric(final_df_to_process['Consecutivo Serie'], errors='coerce')


                            # --- Generación de archivos y guardado (común para ambos modos) ---
                            
                            # Crear columna 'Serie-Número' para el reporte y el guardado
                            s_factura = final_df_to_process['SERIE_FACTURA'].fillna('S/D').astype(str)
                            n_factura = final_df_to_process['NUMERO_FACTURA'].fillna('S/D').astype(str)
                            final_df_to_process['Serie-Número'] = s_factura + "-" + n_factura

                            txt_content = generate_txt_content(final_df_to_process.copy(), account_mappings, tarjetas_destinos)
                            excel_file = generate_excel_report(final_df_to_process.copy())

                            # Preparar datos para guardar en Google Sheets.
                            registros_data_df = final_df_to_process.copy()
                            registros_data_df['Fecha Procesado'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                            # Asegurarse de que todas las columnas esperadas por gspread estén en el DataFrame
                            gsheet_headers = registros_recibos_ws.row_values(1)
                            # Creamos un nuevo DataFrame con las columnas de Google Sheets
                            registros_to_append_df = pd.DataFrame(columns=gsheet_headers)

                            # Mapear columnas del DataFrame a las columnas de Google Sheets.
                            # Nombres en tu DataFrame -> Nombres en Google Sheets
                            col_map = {
                                'SERIE_FACTURA': 'Serie_Factura',
                                'NUMERO_FACTURA': 'Numero_Factura',
                                'Fecha Procesado': 'Fecha Procesado',
                                'Fecha': 'Fecha',
                                'Recibo N°': 'Recibo N°',
                                'Cliente': 'Cliente',
                                'Valor Efectivo': 'Valor Efectivo',
                                'Agrupación': 'Agrupación',
                                'Destino': 'Destino',
                                'Consecutivo Global': 'Consecutivo Global',
                                'Consecutivo Serie': 'Consecutivo Serie',
                                'Serie': 'Serie'
                            }
                            
                            # Rellenar el DataFrame a guardar con los datos del DataFrame final
                            for df_col, gsheet_col in col_map.items():
                                if df_col in registros_data_df.columns and gsheet_col in gsheet_headers:
                                    registros_to_append_df[gsheet_col] = registros_data_df[df_col]
                            
                            # Rellenar las columnas faltantes con cadena vacía.
                            registros_to_append_df = registros_to_append_df[gsheet_headers].fillna('')

                            # Guardar los nuevos (o re-guardados) registros.
                            registros_recibos_ws.append_rows(registros_to_append_df.values.tolist(), value_input_option='USER_ENTERED')
                            
                            st.success("✅ ¡Éxito! Los datos han sido guardados en Google Sheets.")

                            st.subheader("5. Descargar Archivos")
                            dl_col1, dl_col2 = st.columns(2)
                            
                            # Identificador para el nombre de archivo usa el primer (o único) consecutivo global
                            first_global_consecutive = final_df_to_process['Consecutivo Global'].min()
                            file_identifier = f"{serie_seleccionada}_{int(first_global_consecutive)}_{datetime.now().strftime('%Y%m%d')}"
                            
                            with dl_col1:
                                st.download_button(
                                    label="⬇️ Descargar Archivo TXT para el ERP",
                                    data=txt_content.encode('utf-8'),
                                    file_name=f"recibos_{file_identifier}.txt",
                                    mime="text/plain", use_container_width=True
                                )
                            with dl_col2:
                                st.download_button(
                                    label="📄 Descargar Reporte Detallado en Excel",
                                    data=excel_file,
                                    file_name=f"Reporte_Recibos_{file_identifier}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True
                                )

                            # Limpiar sesión para el siguiente ciclo.
                            reset_session_state('mode', 'google_credentials')
                            
                            st.info("El proceso ha finalizado. La página se recargará para iniciar un nuevo ciclo.")
                            time.sleep(5)
                            st.rerun()

                        except Exception as e:
                            st.error(f"Error al guardar los datos o generar los archivos: {e}")


# ==============================================================================
# ================= PESTAÑA 2: IDENTIFICACIÓN (NUEVA FUNCIONALIDAD) ============
# ==============================================================================

with tab2:
    st.header("🏦 Generador de Archivo de Identificación Bancaria")
    st.markdown("""
    Esta herramienta procesa un archivo Excel con dos columnas (`descripcion`, `valor`) y genera el archivo plano para el ERP.
    
    **Reglas de negocio (Actualizadas):**
    1. Se genera una línea de **Crédito (13059901)** y una línea de **Débito (11100502)** por cada registro del Excel.
    2. Serie y Centro de Costo fijos: **189**.
    3. Puedes asignar manualmente el número de serie.
    4. Estructura final: `...|Valor|CentroCosto|NIT|0|0`.
    """)

    # --- INPUTS DE USUARIO ---
    st.divider()
    col_ident_1, col_ident_2, col_ident_3 = st.columns(3)
    with col_ident_1:
        fecha_identificacion = st.date_input("Fecha para el documento:", datetime.now(), key="fecha_ident")
    with col_ident_2:
        consecutivo_ident = st.number_input("Consecutivo del Documento:", value=11899, step=1, key="consec_ident")
    with col_ident_3:
        # Nuevo campo para asignar el número manual después de la serie
        numero_serie_manual = st.text_input("consecutivo manual", value="189", key="num_manual_ident")

    file_ident = st.file_uploader("📂 Cargar Excel (Columnas: descripcion, valor)", type=['xlsx', 'xls'], key="upload_ident")

    def generate_txt_identificacion_custom(df, fecha_str, consecutivo, numero_manual):
        """
        Genera el TXT con formato específico para Identificación.
        Separador: Pipe (|)
        Genera pares de líneas Crédito/Débito por cada fila del Excel.
        """
        lines = []
        
        # Constantes solicitadas
        CTA_CREDITO = "13059901" # Cuenta crédito
        CTA_DEBITO = "11100502"  # Cuenta débito
        TERCERO_NIT = "890903938"      # NIT Bancolombia
        # TERCERO_NOMBRE ELIMINADO SEGUN REQUERIMIENTO
        TIPO_DOC = "10" 
        SERIE = "189"
        CENTRO_COSTO = "189"
        
        total_valor = 0
        
        # Procesar cada fila del Excel
        for _, row in df.iterrows():
            descripcion = str(row['descripcion']).strip()
            # Limpieza básica del valor por si viene con formato moneda o strings
            try:
                val_str_raw = str(row['valor']).replace('$', '').replace(',', '').strip()
                valor = float(val_str_raw)
            except:
                valor = 0.0
            
            total_valor += valor
            valor_formateado = str(round(valor, 2))
            
            # --- LÍNEA 1: CRÉDITO (13059901) ---
            # Estructura Solicitada: 
            # Fecha|Consecutivo|Cuenta|TipoDoc|Detalle|Serie|NumManual|Debito|Credito|CentroCosto|Nit|0|0
            linea_credito = "|".join([
                fecha_str,
                str(consecutivo),
                CTA_CREDITO,
                TIPO_DOC,
                descripcion,
                SERIE,
                str(numero_manual), # Numero asignado por el usuario
                "0",                # Debito
                valor_formateado,   # Credito
                # "0",              # SE ELIMINO LA BASE
                CENTRO_COSTO,       # Centro Costo va despues de los valores
                TERCERO_NIT,        # Luego el NIT
                # SE ELIMINO EL NOMBRE
                "0",                # Filler final 1
                "0"                 # Filler final 2
            ])
            lines.append(linea_credito)

            # --- LÍNEA 2: DÉBITO (11100502) - Contrapartida inmediata ---
            linea_debito = "|".join([
                fecha_str,
                str(consecutivo),
                CTA_DEBITO,
                TIPO_DOC,
                descripcion,
                SERIE,
                str(numero_manual), # Numero asignado por el usuario
                valor_formateado,   # Debito
                "0",                # Credito
                # "0",              # SE ELIMINO LA BASE
                CENTRO_COSTO,       # Centro Costo
                TERCERO_NIT,        # NIT
                # SE ELIMINO EL NOMBRE
                "0",                # Filler final 1
                "0"                 # Filler final 2
            ])
            lines.append(linea_debito)
        
        return "\n".join(lines), total_valor

    if file_ident is not None:
        try:
            df_ident = pd.read_excel(file_ident)
            
            # Normalizar nombres de columnas para ser flexibles (minusculas, sin espacios)
            df_ident.columns = df_ident.columns.astype(str).str.lower().str.strip()
            
            # Validar columnas
            if 'descripcion' in df_ident.columns and 'valor' in df_ident.columns:
                
                # Filtrar filas vacías o con valor 0
                df_ident = df_ident.dropna(subset=['valor'])
                
                st.info("Archivo cargado correctamente. Vista previa de los datos:")
                st.dataframe(df_ident.head(), use_container_width=True)
                
                # Formato fecha DD/MM/AAAA
                fecha_fmt = fecha_identificacion.strftime('%d/%m/%Y')
                
                if st.button("⚙️ Generar TXT de Identificación", type="primary"):
                    txt_content_ident, total_calculado = generate_txt_identificacion_custom(
                        df_ident, 
                        fecha_fmt, 
                        consecutivo_ident,
                        numero_serie_manual # Pasamos el nuevo parametro
                    )
                    
                    st.success(f"¡TXT Generado Exitosamente! Total procesado: ${total_calculado:,.2f}")
                    
                    st.download_button(
                        label="⬇️ Descargar TXT Identificación",
                        data=txt_content_ident.encode('utf-8'),
                        file_name=f"Identificacion_{consecutivo_ident}_{fecha_identificacion.strftime('%Y%m%d')}.txt",
                        mime="text/plain",
                        use_container_width=True
                    )
            else:
                st.error("❌ El archivo de Excel NO tiene las columnas requeridas: 'descripcion' y 'valor'.")
                st.write("Columnas encontradas en tu archivo:", list(df_ident.columns))
                
        except Exception as e:
            st.error(f"Error procesando el archivo: {e}")
