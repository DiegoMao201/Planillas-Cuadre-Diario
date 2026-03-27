# ======================================================================================
# ARCHIVO: 2_Viaticos.py
# VERSIÓN: Módulo de Gestión de Viáticos v1.2 (Excel Profesional)
# ======================================================================================
import streamlit as st
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
import json
import pandas as pd
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app_shared import initialize_access_state, render_sidebar, require_access

# --- 1. CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(layout="wide", page_title="Gestión de Viáticos")

# --- 2. CONEXIÓN A GOOGLE SHEETS (Adaptada para Viáticos) ---
@st.cache_resource(ttl=600)
def connect_to_gsheet_viaticos():
    """Establece conexión con Google Sheets y retorna las hojas para el módulo de viáticos."""
    try:
        creds_json = dict(st.secrets["google_credentials"])
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_json, scope)
        client = gspread.authorize(creds)
        sheet = client.open(st.secrets["google_sheets"]["spreadsheet_name"])
        
        # Hojas de trabajo para Viáticos
        registros_ws = sheet.worksheet("Viaticos_Registros")
        config_ws = sheet.worksheet(st.secrets["google_sheets"]["config_sheet_name"])
        consecutivos_ws = sheet.worksheet("Viaticos_Consecutivos")
        
        return registros_ws, config_ws, consecutivos_ws
    except gspread.exceptions.WorksheetNotFound as e:
        st.error(f"Error fatal: No se encontró la hoja de trabajo '{e.args[0]}'.")
        st.warning("Asegúrese de que las hojas 'Viaticos_Registros' y 'Viaticos_Consecutivos' existan en su Google Sheet.")
        return None, None, None
    except Exception as e:
        st.error(f"Error fatal al conectar con Google Sheets para Viáticos: {e}")
        return None, None, None

# --- 3. LÓGICA DE DATOS Y PROCESAMIENTO ---
def get_viaticos_config(config_ws):
    """Carga la configuración para viáticos: empleados, sedes, categorías de gasto y terceros."""
    try:
        config_data = config_ws.get_all_records()
        
        empleados = sorted(list(set(str(d['Detalle']).strip() for d in config_data if d.get('Tipo Movimiento') == 'EMPLEADO' and d.get('Detalle'))))
        sedes = sorted(list(set(str(d['Sede']).strip() for d in config_data if d.get('Tipo Movimiento') == 'EMPLEADO' and d.get('Sede'))))
        categorias = sorted(list(set(str(d['Detalle']).strip() for d in config_data if d.get('Tipo Movimiento') == 'VIATICO_CATEGORIA' and d.get('Detalle'))))
        terceros = sorted(list(set(str(d['Detalle']).strip() for d in config_data if d.get('Tipo Movimiento') == 'TERCERO' and d.get('Detalle'))))
        
        return empleados, sedes, categorias, terceros
    except Exception as e:
        st.error(f"Error al cargar la configuración de viáticos: {e}")
        return [], [], [], []

def get_account_mappings_viaticos(config_ws):
    """Crea un diccionario de mapeo de cuentas contables para viáticos."""
    try:
        records = config_ws.get_all_records()
        mappings = {}
        for record in records:
            tipo = record.get("Tipo Movimiento")
            detalle = str(record.get("Detalle", "")).strip()
            cuenta = str(record.get("Cuenta Contable", ""))

            if detalle and cuenta:
                if tipo in ["EMPLEADO", "VIATICO_CATEGORIA"]:
                    mappings[detalle] = {'cuenta': cuenta}
                elif tipo == "TERCERO":
                    mappings[detalle] = {
                        'cuenta': cuenta,
                        'nit': str(record.get("NIT", "0")),
                        'nombre': str(record.get("Nombre Tercero", detalle))
                    }
        return mappings
    except Exception as e:
        st.error(f"Error al leer el mapeo de cuentas de viáticos: {e}")
        return {}

def format_currency(num):
    """Formatea un número como moneda colombiana."""
    return f"${int(num):,}".replace(",", ".") if isinstance(num, (int, float)) else "$0"
    
# --- 4. GESTIÓN DEL ESTADO DE LA SESIÓN ---
def initialize_viaticos_state():
    """Inicializa el estado de la sesión para el formulario de viáticos."""
    if 'viaticos_initialized' not in st.session_state:
        st.session_state.viaticos_gastos = []
        st.session_state.viaticos_empleado = None
        st.session_state.viaticos_sede = None
        st.session_state.viaticos_mes = datetime.now().date().replace(day=1)
        st.session_state.viaticos_initialized = True

def clear_viaticos_form():
    """Limpia solo los gastos, manteniendo empleado, sede y mes."""
    st.session_state.viaticos_gastos = []

# --- 5. COMPONENTES DE LA INTERFAZ DE USUARIO (UI) ---
def display_gastos_viaticos_section(categorias_list, terceros_list):
    """Muestra la sección para agregar y editar gastos de viáticos."""
    st.subheader("2. Registro de Gastos", anchor=False, divider="blue")
    
    terceros_con_opciones = ["N/A - Gasto Menor (Doc. Equivalente)", "NUEVO TERCERO (Anexar RUT)"] + terceros_list

    with st.expander("➕ Agregar Nuevo Gasto de Viático", expanded=True):
        with st.form("form_add_gasto_viatico", clear_on_submit=True):
            cols = st.columns([2, 2, 3, 2, 1.5])
            gasto = {
                'Fecha': cols[0].date_input("Fecha Gasto", value=datetime.now().date(), label_visibility="collapsed", format="DD/MM/YYYY"),
                'Categoria': cols[1].selectbox("Categoría", options=categorias_list, label_visibility="collapsed", placeholder="Categoría"),
                'Tercero': cols[2].selectbox("Tercero", options=terceros_con_opciones, label_visibility="collapsed", placeholder="Tercero/Proveedor"),
                'Descripcion': cols[3].text_input("Descripción", label_visibility="collapsed", placeholder="Ej: Peaje La Paila"),
                'Valor': cols[4].number_input("Valor", min_value=1.0, step=1000.0, format="%.0f", label_visibility="collapsed", placeholder="Valor")
            }
            
            if st.form_submit_button("Agregar Gasto", use_container_width=True, type="primary"):
                if gasto['Valor'] > 0 and gasto['Categoria'] and gasto['Tercero'] and gasto['Descripcion']:
                    gasto['Fecha'] = gasto['Fecha'].strftime("%d/%m/%Y")
                    st.session_state.viaticos_gastos.append(gasto)
                    st.toast(f"✅ Gasto de {gasto['Categoria']} por {format_currency(gasto['Valor'])} agregado.")
                    st.rerun()
                else:
                    st.warning("Todos los campos son obligatorios y el valor debe ser mayor a cero.")

    if st.session_state.viaticos_gastos:
        st.markdown("##### Gastos Registrados en este Reporte")
        df = pd.DataFrame(st.session_state.viaticos_gastos)
        df['Eliminar'] = False
        
        column_order = ['Fecha', 'Categoria', 'Tercero', 'Descripcion', 'Valor', 'Eliminar']
        df = df[column_order]

        edited_df = st.data_editor(
            df, key='editor_viaticos', hide_index=True, use_container_width=True,
            column_config={
                "Valor": st.column_config.NumberColumn("Valor", format="$ %.0f", required=True),
                "Eliminar": st.column_config.CheckboxColumn("Eliminar", width="small")
            }
        )
        
        if edited_df['Eliminar'].any():
            indices_to_remove = edited_df[edited_df['Eliminar']].index
            st.session_state.viaticos_gastos = [item for i, item in enumerate(st.session_state.viaticos_gastos) if i not in indices_to_remove]
            st.toast("🗑️ Registro(s) eliminado(s).")
            st.rerun()
        else:
            st.session_state.viaticos_gastos = edited_df.drop(columns=['Eliminar']).to_dict('records')

def display_summary_and_save_viaticos(worksheets):
    """Muestra el resumen de viáticos y el botón para guardar el reporte."""
    st.subheader("3. Verificación y Guardado del Reporte", anchor=False, divider="green")
    
    registros_ws, _, consecutivos_ws = worksheets
    
    with st.container(border=True):
        if not st.session_state.viaticos_gastos:
            st.info("Agregue al menos un gasto para ver el resumen.")
            return
            
        df_gastos = pd.DataFrame(st.session_state.viaticos_gastos)
        total_viaticos = df_gastos['Valor'].sum()
        
        st.metric("💵 **Valor Total del Reporte de Viáticos**", format_currency(total_viaticos))

        st.markdown("##### Resumen por Categoría")
        resumen_cat = df_gastos.groupby('Categoria')['Valor'].sum().reset_index()
        st.dataframe(resumen_cat.style.format({"Valor": format_currency}), use_container_width=True)

        if st.button("💾 Guardar Reporte de Viáticos", type="primary", use_container_width=True):
            empleado = st.session_state.get("viaticos_empleado")
            sede = st.session_state.get("viaticos_sede")
            mes_str = st.session_state.get("viaticos_mes").strftime("%Y-%m")

            if not empleado or not sede:
                st.warning("🛑 Debe seleccionar un empleado y una sede antes de guardar.")
                return

            try:
                cell = consecutivos_ws.find(empleado, in_column=1)
                if cell:
                    next_consecutive = int(consecutivos_ws.cell(cell.row, 2).value) + 1
                    consecutivos_ws.update_cell(cell.row, 2, next_consecutive)
                else:
                    next_consecutive = 1
                    consecutivos_ws.append_row([empleado, next_consecutive])
                
                report_id = f"VT-{empleado.split(' ')[0].upper()}-{mes_str}-{next_consecutive}"
                
                rows_to_add = []
                for gasto in st.session_state.viaticos_gastos:
                    row = [
                        report_id, empleado, sede, mes_str,
                        gasto['Fecha'], gasto['Categoria'], gasto['Tercero'],
                        gasto['Descripcion'], gasto['Valor'],
                        datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                    ]
                    rows_to_add.append(row)
                
                registros_ws.append_rows(rows_to_add)
                
                st.success(f"✅ Reporte de viáticos '{report_id}' guardado con {len(rows_to_add)} gastos.")
                clear_viaticos_form()
            except Exception as e:
                st.error(f"Error al guardar los datos de viáticos: {e}")

# --- 6. GENERACIÓN DE REPORTES (TXT y EXCEL) ---
def generate_excel_report_viaticos(registros_ws, start_date, end_date, selected_employee):
    """Genera un reporte Excel profesional y con formato mejorado de los viáticos."""
    st.info("Generando reporte Excel profesional...")
    try:
        all_records = registros_ws.get_all_records()
        df = pd.DataFrame(all_records)

        if df.empty:
            st.warning("No hay datos en la hoja 'Viaticos_Registros'.")
            return None
            
        df['Valor'] = pd.to_numeric(df['Valor'])
        df['Fecha_Gasto_dt'] = pd.to_datetime(df['Fecha_Gasto'], format='%d/%m/%Y')

        mask = (df['Fecha_Gasto_dt'].dt.date >= start_date) & (df['Fecha_Gasto_dt'].dt.date <= end_date)
        if selected_employee != "Todos los Empleados":
            mask &= (df['Empleado'] == selected_employee)
        
        filtered_df = df[mask].sort_values(by=['Empleado', 'Reporte_ID', 'Fecha_Gasto_dt'])

        if filtered_df.empty:
            st.warning("No se encontraron registros de viáticos para los filtros seleccionados.")
            return None
        
        output = io.BytesIO()
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Reporte de Viáticos"

        # --- Estilos Profesionales ---
        font_title = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
        fill_title = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        font_header = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font_group_header = Font(name='Calibri', size=12, bold=True)
        fill_group_header = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        font_total = Font(name='Calibri', size=11, bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        currency_format = '$ #,##0'
        date_format = 'DD/MM/YYYY'

        # --- Título Principal ---
        ws.merge_cells('A1:H2')
        title_cell = ws['A1']
        title_cell.value = "REPORTE DETALLADO DE VIÁTICOS"
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = align_center
        ws['A3'] = f"Período del {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}"
        ws.merge_cells('A3:H3')
        ws['A3'].alignment = align_center
        ws['A3'].font = Font(italic=True)

        current_row = 5
        # --- Encabezados de la tabla ---
        headers = ["Reporte ID", "Empleado", "Sede", "Fecha Gasto", "Categoría", "Tercero", "Descripción", "Valor"]
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header_title)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = thin_border
            cell.alignment = align_center
        
        current_row += 1
        
        # --- Escribir datos agrupados por reporte ---
        grand_total = 0
        for report_id, group in filtered_df.groupby('Reporte_ID'):
            report_total = group['Valor'].sum()
            grand_total += report_total
            
            # Fila de cabecera para el grupo
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            group_header_cell = ws.cell(row=current_row, column=1, value=f"Detalle del Reporte: {report_id}  (Empleado: {group['Empleado'].iloc[0]})")
            group_header_cell.font = font_group_header
            group_header_cell.fill = fill_group_header
            
            # Celda para el total del grupo
            total_group_cell = ws.cell(row=current_row, column=8)
            total_group_cell.font = font_group_header
            total_group_cell.fill = fill_group_header
            
            current_row += 1

            for _, row_data in group.iterrows():
                ws.cell(row=current_row, column=1, value=row_data['Reporte_ID'])
                ws.cell(row=current_row, column=2, value=row_data['Empleado'])
                ws.cell(row=current_row, column=3, value=row_data['Sede'])
                date_cell = ws.cell(row=current_row, column=4, value=row_data['Fecha_Gasto_dt'])
                date_cell.number_format = date_format
                ws.cell(row=current_row, column=5, value=row_data['Categoria'])
                ws.cell(row=current_row, column=6, value=row_data['Tercero'])
                ws.cell(row=current_row, column=7, value=row_data['Descripcion'])
                value_cell = ws.cell(row=current_row, column=8, value=row_data['Valor'])
                value_cell.number_format = currency_format
                value_cell.alignment = align_right
                current_row += 1

            # Escribir el total del reporte
            total_group_cell.value = report_total
            total_group_cell.number_format = currency_format
            total_group_cell.alignment = align_right

        # --- Gran Total ---
        ws.cell(row=current_row, column=7, value="GRAN TOTAL").font = font_total
        ws.cell(row=current_row, column=7, value="GRAN TOTAL").alignment = align_right
        total_cell = ws.cell(row=current_row, column=8, value=grand_total)
        total_cell.font = font_total
        total_cell.number_format = currency_format
        total_cell.alignment = align_right

        # --- Ajustar Ancho de Columnas ---
        column_widths = {'A': 20, 'B': 25, 'C': 15, 'D': 15, 'E': 18, 'F': 25, 'G': 35, 'H': 18}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        st.error(f"Error al generar el reporte Excel de viáticos: {e}")
        return None

def generate_txt_file_viaticos(registros_ws, config_ws, start_date, end_date, selected_employee):
    """Genera el archivo TXT para el ERP con los datos de viáticos."""
    st.info("Generando archivo TXT para contabilidad...")
    
    try:
        all_records = registros_ws.get_all_records()
        account_mappings = get_account_mappings_viaticos(config_ws)

        if not account_mappings:
            st.error("No se pudo generar el TXT: Faltan mapeos de cuentas en 'Configuracion'.")
            return None
        
        df = pd.DataFrame(all_records)
        if df.empty:
            st.warning("No se encontraron registros para generar el archivo TXT.")
            return None
            
        df['Valor'] = pd.to_numeric(df['Valor'])
        df['Fecha_Gasto_dt'] = pd.to_datetime(df['Fecha_Gasto'], format='%d/%m/%Y')

        mask = (df['Fecha_Gasto_dt'].dt.date >= start_date) & (df['Fecha_Gasto_dt'].dt.date <= end_date)
        if selected_employee != "Todos los Empleados":
            mask &= (df['Empleado'] == selected_employee)
        
        filtered_records = df[mask]

        if filtered_records.empty:
            st.warning("No se encontraron registros para generar el archivo TXT.")
            return None

        txt_lines = []
        for report_id, group in filtered_records.groupby('Reporte_ID'):
            total_reporte = group['Valor'].sum()
            fecha_reporte = group['Fecha_Gasto_dt'].max().strftime('%d/%m/%Y')
            empleado = group['Empleado'].iloc[0]
            sede = group['Sede'].iloc[0]
            
            for _, row in group.iterrows():
                categoria_gasto = row['Categoria']
                tercero_gasto = row['Tercero']
                
                cuenta_debito = account_mappings.get(categoria_gasto, {}).get('cuenta', f'ERR_{categoria_gasto}')
                tercero_info = account_mappings.get(tercero_gasto, {})
                nit_tercero = tercero_info.get('nit', '0')

                linea_debito = "|".join([
                    str(fecha_reporte), str(report_id), str(cuenta_debito), "10",
                    f"Viatico {row['Descripcion']}", str(sede), str(report_id),
                    str(row['Valor']), "0", str(sede), str(nit_tercero), "0", "0"
                ])
                txt_lines.append(linea_debito)

            cuenta_credito_empleado = account_mappings.get(empleado, {}).get('cuenta', f'ERR_{empleado}')
            linea_credito = "|".join([
                str(fecha_reporte), str(report_id), str(cuenta_credito_empleado), "10",
                f"Causación Viáticos {empleado} - Reporte {report_id}", str(sede), str(report_id),
                "0", str(total_reporte), str(sede), "0", "0", "0"
            ])
            txt_lines.append(linea_credito)
            
        return "\n".join(txt_lines)

    except Exception as e:
        st.error(f"Error crítico al generar el archivo TXT: {e}")
        return None

# --- 7. FLUJO PRINCIPAL DE LA APLICACIÓN ---
def main():
    """Función principal que ejecuta la aplicación de Viáticos."""
    st.title("✈️ Módulo de Gestión de Viáticos")

    worksheets = connect_to_gsheet_viaticos()
    
    if all(worksheets):
        registros_ws, config_ws, _ = worksheets
        
        config_data = get_viaticos_config(config_ws)
        empleados, sedes, categorias, terceros = config_data

        if not empleados or not categorias:
            st.error("🚨 Faltan datos en la hoja 'Configuracion'.")
            st.warning("Asegúrese de haber definido al menos un 'EMPLEADO' y una 'VIATICO_CATEGORIA'.")
            return

        tab_form, tab_reports = st.tabs(["📝 Registrar Reporte", "📈 Generar Reportes"])

        with tab_form:
            st.header("Formulario de Registro de Viáticos", anchor=False)
            st.subheader("1. Información del Reporte", anchor=False, divider="red")

            col1, col2, col3, col4 = st.columns([2, 2, 1, 1])
            st.session_state.viaticos_empleado = col1.selectbox("Empleado", options=empleados, key="sb_empleado", placeholder="Seleccione empleado...")
            st.session_state.viaticos_sede = col2.selectbox("Sede de Trabajo", options=sedes, key="sb_sede", placeholder="Seleccione sede...")
            
            current_year = datetime.now().year
            current_month = datetime.now().month
            
            selected_year = col3.selectbox(
                "Año", options=range(current_year + 1, current_year - 5, -1), key="sb_year"
            )
            selected_month = col4.selectbox(
                "Mes", options=range(1, 13), 
                format_func=lambda month: datetime(current_year, month, 1).strftime("%B"),
                index=current_month - 1, key="sb_month"
            )
            
            st.session_state.viaticos_mes = datetime(selected_year, selected_month, 1).date()

            if st.button("✨ Iniciar Nuevo Reporte (limpiar gastos)", use_container_width=True):
                clear_viaticos_form()
                st.rerun()
            
            st.divider()
            
            display_gastos_viaticos_section(categorias, terceros)
            display_summary_and_save_viaticos(worksheets)

        with tab_reports:
            st.header("Generación de Archivos y Reportes de Viáticos", anchor=False)
            
            today = datetime.now().date()
            rep_col1, rep_col2, rep_col3 = st.columns(3)
            
            employee_options = ["Todos los Empleados"] + empleados
            selected_employee_rep = rep_col1.selectbox("Filtrar por Empleado", options=employee_options, key="sb_rep_emp")
            start_date_rep = rep_col2.date_input("Fecha de Inicio", today.replace(day=1), key="di_rep_start")
            end_date_rep = rep_col3.date_input("Fecha de Fin", today, key="di_rep_end")

            if start_date_rep > end_date_rep:
                st.error("Error: La fecha de inicio no puede ser posterior a la fecha de fin.")
            else:
                st.divider()
                b1, b2 = st.columns(2)
                
                with b1:
                    if st.button("📄 Generar Archivo TXT para ERP", use_container_width=True, type="primary"):
                        txt_content = generate_txt_file_viaticos(registros_ws, config_ws, start_date_rep, end_date_rep, selected_employee_rep)
                        if txt_content:
                            st.download_button(
                                label="📥 Descargar .txt de Viáticos",
                                data=txt_content.encode('utf-8'),
                                file_name=f"viaticos_{start_date_rep.strftime('%Y%m%d')}_{end_date_rep.strftime('%Y%m%d')}.txt",
                                mime="text/plain",
                                use_container_width=True
                            )
                
                with b2:
                    if st.button("📊 Generar Reporte Detallado en Excel", use_container_width=True, type="primary"):
                        excel_data = generate_excel_report_viaticos(registros_ws, start_date_rep, end_date_rep, selected_employee_rep)
                        if excel_data:
                            st.download_button(
                                label="📥 Descargar .xlsx de Viáticos",
                                data=excel_data,
                                file_name=f"Reporte_Viaticos_{selected_employee_rep.replace(' ','_')}_{start_date_rep.strftime('%Y%m%d')}_{end_date_rep.strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
    else:
        st.info("⏳ Esperando conexión con Google Sheets...")

# --- BLOQUE DE EJECUCIÓN PRINCIPAL ---
if __name__ == "__main__":
    initialize_access_state()
    require_access(
        "admin",
        "Gestion de viaticos",
        "Este modulo queda bajo acceso administrativo.",
    )
    render_sidebar("Viaticos")
    initialize_viaticos_state()
    main()
