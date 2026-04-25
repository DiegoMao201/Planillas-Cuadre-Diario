from io import BytesIO
from datetime import date, datetime

import pandas as pd
import streamlit as st
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app_shared import (
    AUDIT_HEADERS,
    LOGO_PATH,
    NOVEDADES_HEADERS,
    append_audit_log,
    append_novedad,
    build_whatsapp_url,
    current_colombia_datetime,
    current_colombia_date,
    format_colombia_timestamp,
    get_auxiliary_records,
    get_request_records,
    load_solicitudes_management_data,
    get_solicitudes_worksheets,
    inject_shared_css,
    initialize_access_state,
    refresh_management_report,
    render_brand_header,
    render_sidebar,
    require_access,
    send_employee_response_email,
    update_request_record,
)


st.set_page_config(layout="wide", page_title="Gestion de Solicitudes")

NAVY = "0B3954"
ORANGE = "E67E22"
TEAL = "1F7A8C"
LIGHT = "F4F8FB"
MID = "D7E3EB"
TEXT = "183B56"


def inject_management_css() -> None:
    st.markdown(
        f"""
        <style>
        .fx-hero {{
            background: linear-gradient(135deg, #{NAVY} 0%, #1d5d82 62%, #{TEAL} 100%);
            border-radius: 24px;
            padding: 1.5rem 1.6rem;
            color: #ffffff;
            box-shadow: 0 18px 38px rgba(11, 57, 84, 0.18);
            margin-bottom: 1rem;
        }}
        .fx-hero h2 {{
            margin: 0;
            font-size: 1.45rem;
            font-weight: 800;
            letter-spacing: 0.01em;
        }}
        .fx-hero p {{
            margin: 0.45rem 0 0 0;
            max-width: 900px;
            color: rgba(255, 255, 255, 0.88);
        }}
        .fx-chip-row {{
            display: flex;
            flex-wrap: wrap;
            gap: 0.5rem;
            margin-top: 1rem;
        }}
        .fx-chip {{
            background: rgba(255, 255, 255, 0.14);
            border: 1px solid rgba(255, 255, 255, 0.18);
            border-radius: 999px;
            padding: 0.38rem 0.8rem;
            font-size: 0.84rem;
            font-weight: 600;
        }}
        .fx-kpi-card {{
            border-radius: 20px;
            padding: 1rem 1.1rem;
            background: #ffffff;
            border: 1px solid rgba(11, 57, 84, 0.08);
            box-shadow: 0 12px 26px rgba(11, 57, 84, 0.06);
            min-height: 145px;
        }}
        .fx-kpi-card .label {{
            color: #557085;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            font-size: 0.76rem;
            font-weight: 800;
        }}
        .fx-kpi-card .value {{
            color: #{TEXT};
            font-size: 2rem;
            line-height: 1.05;
            font-weight: 800;
            margin-top: 0.45rem;
        }}
        .fx-kpi-card .note {{
            color: #678196;
            font-size: 0.86rem;
            margin-top: 0.45rem;
        }}
        .fx-kpi-card.navy {{ border-top: 5px solid #{NAVY}; }}
        .fx-kpi-card.orange {{ border-top: 5px solid #{ORANGE}; }}
        .fx-kpi-card.teal {{ border-top: 5px solid #{TEAL}; }}
        .fx-kpi-card.slate {{ border-top: 5px solid #6C7F90; }}
        .fx-panel {{
            border-radius: 22px;
            padding: 1.2rem 1.25rem;
            background: linear-gradient(180deg, #ffffff 0%, #{LIGHT} 100%);
            border: 1px solid rgba(11, 57, 84, 0.08);
            box-shadow: 0 12px 30px rgba(11, 57, 84, 0.06);
        }}
        .fx-panel h3 {{
            margin: 0 0 0.2rem 0;
            color: #{TEXT};
            font-size: 1.02rem;
        }}
        .fx-panel p {{
            margin: 0;
            color: #678196;
            font-size: 0.88rem;
        }}
        .fx-mini-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 0.75rem;
            margin-top: 0.95rem;
        }}
        .fx-mini-card {{
            border-radius: 16px;
            padding: 0.85rem 0.9rem;
            background: #ffffff;
            border: 1px solid rgba(11, 57, 84, 0.07);
        }}
        .fx-mini-card .mini-label {{
            color: #6d8598;
            text-transform: uppercase;
            font-size: 0.73rem;
            font-weight: 800;
            letter-spacing: 0.04em;
        }}
        .fx-mini-card .mini-value {{
            color: #{TEXT};
            font-size: 1rem;
            font-weight: 700;
            margin-top: 0.3rem;
        }}
        .fx-divider-space {{
            height: 0.35rem;
        }}
        .stTabs [data-baseweb="tab-list"] {{
            gap: 0.35rem;
        }}
        .stTabs [data-baseweb="tab"] {{
            border-radius: 14px 14px 0 0;
            padding: 0.75rem 1rem;
            background: #eef4f8;
            border: 1px solid rgba(11, 57, 84, 0.08);
            color: #{TEXT};
            font-weight: 700;
        }}
        .stTabs [aria-selected="true"] {{
            background: #ffffff !important;
            border-bottom-color: #ffffff !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def _safe_percentage(value: int, total: int) -> str:
    if not total:
        return "0%"
    return f"{(value / total) * 100:.1f}%"


def _parse_datetime(value: pd.Series) -> pd.Series:
    return pd.to_datetime(value, format="%d/%m/%Y %H:%M:%S", errors="coerce")


def prepare_request_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    working_df = df.copy()
    working_df["Estado"] = working_df["Estado"].replace("", "Pendiente").fillna("Pendiente")
    working_df["Tipo_Solicitud"] = working_df["Tipo_Solicitud"].replace("", "Sin tipo").fillna("Sin tipo")
    working_df["Sede"] = working_df["Sede"].replace("", "Sin sede").fillna("Sin sede")
    working_df["Nombre_Completo"] = working_df["Nombre_Completo"].replace("", "Sin nombre").fillna("Sin nombre")
    working_df["Fecha_Solicitud_dt"] = pd.to_datetime(working_df["Fecha_Solicitud"], format="%d/%m/%Y", errors="coerce")
    working_df["Fecha_Inicial_dt"] = pd.to_datetime(working_df["Fecha_Inicial"], format="%d/%m/%Y", errors="coerce")
    working_df["Fecha_Final_dt"] = pd.to_datetime(working_df["Fecha_Final"], format="%d/%m/%Y", errors="coerce")
    working_df["Fecha_Registro_dt"] = _parse_datetime(working_df["Fecha_Registro"])
    working_df["Fecha_Respuesta_dt"] = _parse_datetime(working_df["Fecha_Respuesta"])
    reference_dates = working_df["Fecha_Registro_dt"].fillna(working_df["Fecha_Solicitud_dt"])
    working_df["Tiempo_Resolucion_Horas"] = (
        working_df["Fecha_Respuesta_dt"] - reference_dates
    ).dt.total_seconds() / 3600
    working_df["Periodo_Mensual"] = working_df["Fecha_Solicitud_dt"].dt.strftime("%Y-%m").fillna("Sin fecha")
    working_df.sort_values(by=["Fecha_Solicitud_dt", "Fecha_Registro_dt"], ascending=[False, False], inplace=True)
    return working_df


def _date_bounds(df: pd.DataFrame) -> tuple[date, date]:
    valid_dates = df["Fecha_Solicitud_dt"].dropna()
    if valid_dates.empty:
        today = current_colombia_date()
        return today, today
    return valid_dates.min().date(), valid_dates.max().date()


def vacations_snapshot(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    today = current_colombia_date()
    vacations_df = df[(df["Tipo_Solicitud"] == "Vacaciones") & (df["Estado"] == "Aprobada")].copy()
    active_df = vacations_df[
        (vacations_df["Fecha_Inicial_dt"].dt.date <= today)
        & (vacations_df["Fecha_Final_dt"].dt.date >= today)
    ].copy()
    upcoming_df = vacations_df[
        vacations_df["Fecha_Inicial_dt"].dt.date > today
    ].copy()
    upcoming_df.sort_values(by=["Fecha_Inicial_dt", "Sede", "Nombre_Completo"], inplace=True)
    active_df.sort_values(by=["Sede", "Nombre_Completo", "Fecha_Inicial_dt"], inplace=True)
    return active_df, upcoming_df


def group_summary(df: pd.DataFrame, column_name: str, empty_label: str, top_n: int | None = None) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=[column_name, "Cantidad", "Participacion"])

    summary = (
        df.assign(**{column_name: df[column_name].replace("", empty_label).fillna(empty_label)})
        .groupby(column_name)
        .size()
        .reset_index(name="Cantidad")
        .sort_values(by=["Cantidad", column_name], ascending=[False, True])
    )
    total = int(summary["Cantidad"].sum())
    summary["Participacion"] = summary["Cantidad"].apply(lambda value: _safe_percentage(int(value), total))
    if top_n:
        return summary.head(top_n)
    return summary


def render_kpi_card(label: str, value: str, note: str, tone: str) -> None:
    st.markdown(
        f"""
        <div class="fx-kpi-card {tone}">
            <div class="label">{label}</div>
            <div class="value">{value}</div>
            <div class="note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_info_panel(title: str, subtitle: str, content_html: str) -> None:
    st.markdown(
        f"""
        <div class="fx-panel">
            <h3>{title}</h3>
            <p>{subtitle}</p>
            <div class="fx-divider-space"></div>
            {content_html}
        </div>
        """,
        unsafe_allow_html=True,
    )


def build_filter_summary(start_date: date, end_date: date, status_filter: str, type_filter: str, sede_filter: str, cedula_filter: str, name_filter: str) -> dict[str, str]:
    return {
        "Rango de fechas": f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}",
        "Estado": status_filter,
        "Tipo": type_filter,
        "Sede": sede_filter,
        "Cedula": cedula_filter.strip() or "Todas",
        "Empleado": name_filter.strip() or "Todos",
    }


def style_sheet(worksheet, header_row: int, freeze_cell: str) -> None:
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color=MID),
        right=Side(style="thin", color=MID),
        top=Side(style="thin", color=MID),
        bottom=Side(style="thin", color=MID),
    )

    worksheet.freeze_panes = freeze_cell
    for cell in worksheet[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in worksheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 4, 34)


def generate_report_excel(
    report_df: pd.DataFrame,
    novedades_df: pd.DataFrame,
    audit_df: pd.DataFrame,
    filter_summary: dict[str, str],
) -> bytes:
    output = BytesIO()
    request_ids = report_df["Solicitud_ID"].tolist() if "Solicitud_ID" in report_df.columns else []

    detail_columns = [
        "Solicitud_ID",
        "Fecha_Registro",
        "Fecha_Solicitud",
        "Estado",
        "Tipo_Solicitud",
        "Motivo_Descripcion",
        "Cedula",
        "Nombre_Completo",
        "Cargo",
        "Sede",
        "Fecha_Inicial",
        "Fecha_Final",
        "Hora_Salida",
        "Tiempo_Total",
        "Persona_A_Cargo",
        "Dias_Aprobados",
        "Periodo_Correspondiente",
        "Dias_Pendientes_Periodo",
        "Fecha_Reincorporacion",
        "Responsable_Revision",
        "Fecha_Respuesta",
        "Medio_Respuesta",
        "Observaciones_RRHH",
        "Correo_Empleado",
        "Telefono_Empleado",
        "Adjunto_URL",
        "Ultima_Actualizacion",
    ]
    rename_map = {
        "Solicitud_ID": "Solicitud ID",
        "Fecha_Registro": "Fecha Registro",
        "Fecha_Solicitud": "Fecha Solicitud",
        "Estado": "Estado",
        "Tipo_Solicitud": "Tipo Solicitud",
        "Motivo_Descripcion": "Motivo",
        "Cedula": "Cedula",
        "Nombre_Completo": "Nombre Completo",
        "Cargo": "Cargo",
        "Sede": "Sede",
        "Fecha_Inicial": "Fecha Inicial",
        "Fecha_Final": "Fecha Final",
        "Hora_Salida": "Hora Salida",
        "Tiempo_Total": "Tiempo Total",
        "Persona_A_Cargo": "Persona a Cargo",
        "Dias_Aprobados": "Dias Aprobados",
        "Periodo_Correspondiente": "Periodo Correspondiente",
        "Dias_Pendientes_Periodo": "Dias Pendientes Periodo",
        "Fecha_Reincorporacion": "Fecha Reincorporacion",
        "Responsable_Revision": "Responsable Revision",
        "Fecha_Respuesta": "Fecha Respuesta",
        "Medio_Respuesta": "Medio Respuesta",
        "Observaciones_RRHH": "Observaciones RRHH",
        "Correo_Empleado": "Correo Empleado",
        "Telefono_Empleado": "Telefono Empleado",
        "Adjunto_URL": "Adjunto URL",
        "Ultima_Actualizacion": "Ultima Actualizacion",
    }

    export_df = report_df.copy()
    for column in detail_columns:
        if column not in export_df.columns:
            export_df[column] = ""
    export_df = export_df[detail_columns].rename(columns=rename_map)

    novedades_export = novedades_df[novedades_df["Solicitud_ID"].isin(request_ids)].copy() if not novedades_df.empty else novedades_df.copy()
    audit_export = audit_df[audit_df["Solicitud_ID"].isin(request_ids)].copy() if not audit_df.empty else audit_df.copy()

    status_summary = group_summary(report_df, "Estado", "Pendiente")
    type_summary = group_summary(report_df, "Tipo_Solicitud", "Sin tipo")
    sede_summary = group_summary(report_df, "Sede", "Sin sede")

    resolved = report_df[report_df["Estado"].isin(["Aprobada", "Negada"])]
    avg_response_hours = resolved["Tiempo_Resolucion_Horas"].dropna().mean() if "Tiempo_Resolucion_Horas" in resolved.columns else 0
    avg_response_display = f"{avg_response_hours:.1f} horas" if pd.notna(avg_response_hours) else "Sin dato"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Detalle Solicitudes", startrow=4)
        novedades_export.to_excel(writer, index=False, sheet_name="Novedades", startrow=2)
        audit_export.to_excel(writer, index=False, sheet_name="Auditoria", startrow=2)

        workbook = writer.book
        summary_ws = workbook.create_sheet("Resumen Ejecutivo", 0)

        summary_ws.merge_cells("B2:G2")
        summary_ws["B2"] = "FERREINOX | Reporte Gerencial de Solicitudes"
        summary_ws["B2"].font = Font(size=16, bold=True, color=TEXT)
        summary_ws["B3"] = "Corte generado"
        summary_ws["C3"] = format_colombia_timestamp()
        summary_ws["B4"] = "Origen"
        summary_ws["C4"] = "Portal administrativo de solicitudes"

        if LOGO_PATH.exists():
            try:
                logo = XLImage(str(LOGO_PATH))
                logo.width = 155
                logo.height = 55
                summary_ws.add_image(logo, "H2")
            except Exception:
                pass

        summary_ws["B6"] = "Filtros aplicados"
        summary_ws["B6"].font = Font(size=12, bold=True, color=TEXT)
        current_row = 7
        for label, value in filter_summary.items():
            summary_ws[f"B{current_row}"] = label
            summary_ws[f"C{current_row}"] = value
            current_row += 1

        current_row += 1
        summary_ws[f"B{current_row}"] = "Indicadores ejecutivos"
        summary_ws[f"B{current_row}"].font = Font(size=12, bold=True, color=TEXT)
        current_row += 1
        metrics = [
            ("Solicitudes en corte", len(report_df)),
            ("Pendientes", int((report_df["Estado"] == "Pendiente").sum())),
            ("En revision", int((report_df["Estado"] == "En revision").sum())),
            ("Aprobadas", int((report_df["Estado"] == "Aprobada").sum())),
            ("Negadas", int((report_df["Estado"] == "Negada").sum())),
            ("Tiempo promedio de respuesta", avg_response_display),
        ]
        for label, value in metrics:
            summary_ws[f"B{current_row}"] = label
            summary_ws[f"C{current_row}"] = value
            current_row += 1

        def write_summary_table(start_col: str, start_row: int, title: str, summary_df: pd.DataFrame, column_name: str) -> None:
            summary_ws[f"{start_col}{start_row}"] = title
            summary_ws[f"{start_col}{start_row}"].font = Font(size=12, bold=True, color=TEXT)
            header_row = start_row + 1
            columns = [column_name, "Cantidad", "Participacion"]
            for offset, header in enumerate(columns):
                cell = summary_ws.cell(row=header_row, column=summary_ws[f"{start_col}1"].column + offset)
                cell.value = header
                cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            row_pointer = header_row + 1
            for _, row in summary_df.iterrows():
                summary_ws.cell(row=row_pointer, column=summary_ws[f"{start_col}1"].column, value=row[column_name])
                summary_ws.cell(row=row_pointer, column=summary_ws[f"{start_col}1"].column + 1, value=int(row["Cantidad"]))
                summary_ws.cell(row=row_pointer, column=summary_ws[f"{start_col}1"].column + 2, value=row["Participacion"])
                row_pointer += 1

        write_summary_table("E", 6, "Estado", status_summary, "Estado")
        write_summary_table("I", 6, "Tipo de solicitud", type_summary, "Tipo_Solicitud")
        write_summary_table("M", 6, "Sede", sede_summary, "Sede")

        for column in range(2, 16):
            summary_ws.column_dimensions[get_column_letter(column)].width = 24

        detail_ws = writer.sheets["Detalle Solicitudes"]
        detail_ws["A1"] = "FERREINOX | Detalle profesional de solicitudes"
        detail_ws["A1"].font = Font(size=14, bold=True, color=TEXT)
        detail_ws["A2"] = "Rango exportado"
        detail_ws["B2"] = filter_summary["Rango de fechas"]
        detail_ws["A3"] = "Generado"
        detail_ws["B3"] = format_colombia_timestamp()
        style_sheet(detail_ws, header_row=5, freeze_cell="A6")

        novedades_ws = writer.sheets["Novedades"]
        novedades_ws["A1"] = "FERREINOX | Novedades asociadas al corte"
        novedades_ws["A1"].font = Font(size=13, bold=True, color=TEXT)
        if novedades_export.empty:
            novedades_ws["A3"] = "No se encontraron novedades para el corte seleccionado."
        else:
            style_sheet(novedades_ws, header_row=3, freeze_cell="A4")

        audit_ws = writer.sheets["Auditoria"]
        audit_ws["A1"] = "FERREINOX | Auditoria del corte"
        audit_ws["A1"].font = Font(size=13, bold=True, color=TEXT)
        if audit_export.empty:
            audit_ws["A3"] = "No se encontraron eventos de auditoria para el corte seleccionado."
        else:
            style_sheet(audit_ws, header_row=3, freeze_cell="A4")

    output.seek(0)
    return output.getvalue()


def main() -> None:
    initialize_access_state()
    require_access(
        "admin",
        "Gestion administrativa de solicitudes",
        "Vista ejecutiva para aprobacion, seguimiento, reportes gerenciales y respuesta al colaborador.",
    )
    inject_shared_css()
    inject_management_css()
    render_sidebar("Gestion de solicitudes")
    render_brand_header(
        "Centro Gerencial de Solicitudes",
        "Panel ejecutivo para decisiones de talento humano, control por fechas y exportacion profesional de informacion.",
    )

    worksheets = get_solicitudes_worksheets()
    management_data = load_solicitudes_management_data()
    df = prepare_request_dataframe(management_data["requests"])
    novedades_df = management_data["novedades"]
    audit_df = management_data["audit"]
    refresh_management_report(worksheets["reporte"], df)

    if df.empty:
        st.info("Aun no existen solicitudes registradas.")
        return

    min_date, max_date = _date_bounds(df)
    st.markdown(
        f"""
        <div class="fx-hero">
            <h2>Reporte ejecutivo para aprobacion, control y trazabilidad</h2>
            <p>
                Esta vista consolida indicadores, volumen por sede, comportamiento por tipo de solicitud,
                capacidad de respuesta y gestion operativa en una sola experiencia administrativa.
            </p>
            <div class="fx-chip-row">
                <span class="fx-chip">Corte disponible: {min_date.strftime('%d/%m/%Y')} a {max_date.strftime('%d/%m/%Y')}</span>
                <span class="fx-chip">Actualizacion automatica de la hoja gerencial en Google Sheets</span>
                <span class="fx-chip">Uso exclusivo administrador</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    filter_cols = st.columns([1.4, 1, 1, 1, 1, 1])
    selected_range = filter_cols[0].date_input(
        "Rango de fechas de solicitud",
        value=(min_date, max_date),
        min_value=min_date,
        max_value=max_date,
    )
    if isinstance(selected_range, tuple) and len(selected_range) == 2:
        start_date, end_date = selected_range
    elif isinstance(selected_range, list) and len(selected_range) == 2:
        start_date, end_date = selected_range[0], selected_range[1]
    else:
        start_date = end_date = selected_range

    status_options = ["Todos"] + sorted(df["Estado"].unique().tolist())
    type_options = ["Todos"] + sorted(df["Tipo_Solicitud"].unique().tolist())
    sede_options = ["Todas"] + sorted(df["Sede"].unique().tolist())
    status_filter = filter_cols[1].selectbox("Estado", status_options)
    type_filter = filter_cols[2].selectbox("Tipo", type_options)
    sede_filter = filter_cols[3].selectbox("Sede", sede_options)
    cedula_filter = filter_cols[4].text_input("Cedula", placeholder="Todas")
    name_filter = filter_cols[5].text_input("Empleado", placeholder="Todos")

    filtered_df = df[
        (df["Fecha_Solicitud_dt"].dt.date >= start_date)
        & (df["Fecha_Solicitud_dt"].dt.date <= end_date)
    ].copy()
    if status_filter != "Todos":
        filtered_df = filtered_df[filtered_df["Estado"] == status_filter]
    if type_filter != "Todos":
        filtered_df = filtered_df[filtered_df["Tipo_Solicitud"] == type_filter]
    if sede_filter != "Todas":
        filtered_df = filtered_df[filtered_df["Sede"] == sede_filter]
    if cedula_filter.strip():
        filtered_df = filtered_df[filtered_df["Cedula"].astype(str).str.contains(cedula_filter.strip(), case=False, na=False)]
    if name_filter.strip():
        filtered_df = filtered_df[filtered_df["Nombre_Completo"].astype(str).str.contains(name_filter.strip(), case=False, na=False)]

    if filtered_df.empty:
        st.warning("No hay solicitudes para el corte y filtros seleccionados.")
        return

    total_requests = len(filtered_df)
    pending_count = int((filtered_df["Estado"] == "Pendiente").sum())
    review_count = int((filtered_df["Estado"] == "En revision").sum())
    approved_count = int((filtered_df["Estado"] == "Aprobada").sum())
    denied_count = int((filtered_df["Estado"] == "Negada").sum())
    resolved_df = filtered_df[filtered_df["Estado"].isin(["Aprobada", "Negada"])]
    avg_response_hours = resolved_df["Tiempo_Resolucion_Horas"].dropna().mean()
    avg_response_display = f"{avg_response_hours:.1f} h promedio" if pd.notna(avg_response_hours) else "Sin historial suficiente"

    kpi_cols = st.columns(5)
    with kpi_cols[0]:
        render_kpi_card("Solicitudes en corte", str(total_requests), f"Periodo {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}", "navy")
    with kpi_cols[1]:
        render_kpi_card("Pendientes", str(pending_count), f"{_safe_percentage(pending_count, total_requests)} del total", "orange")
    with kpi_cols[2]:
        render_kpi_card("En revision", str(review_count), "Casos que requieren seguimiento inmediato", "slate")
    with kpi_cols[3]:
        render_kpi_card("Aprobadas", str(approved_count), f"{_safe_percentage(approved_count, total_requests)} del total", "teal")
    with kpi_cols[4]:
        render_kpi_card("Tiempo de respuesta", avg_response_display, "Calculado para solicitudes con cierre registrado", "navy")

    filter_summary = build_filter_summary(
        start_date,
        end_date,
        status_filter,
        type_filter,
        sede_filter,
        cedula_filter,
        name_filter,
    )
    report_excel = generate_report_excel(filtered_df, novedades_df, audit_df, filter_summary)

    status_summary = group_summary(filtered_df, "Estado", "Pendiente")
    type_summary = group_summary(filtered_df, "Tipo_Solicitud", "Sin tipo", top_n=8)
    sede_summary = group_summary(filtered_df, "Sede", "Sin sede", top_n=8)
    employee_summary = group_summary(filtered_df, "Nombre_Completo", "Sin nombre", top_n=12)
    monthly_summary = (
        filtered_df.groupby("Periodo_Mensual")
        .size()
        .reset_index(name="Cantidad")
        .sort_values(by="Periodo_Mensual")
    )
    active_vacations_df, upcoming_vacations_df = vacations_snapshot(filtered_df)
    vacations_by_sede = group_summary(active_vacations_df, "Sede", "Sin sede") if not active_vacations_df.empty else pd.DataFrame(columns=["Sede", "Cantidad", "Participacion"])
    today_colombia = current_colombia_date()
    pending_focus_df = filtered_df[filtered_df["Estado"].isin(["Pendiente", "En revision"])].copy()
    pending_focus_df.sort_values(by=["Fecha_Solicitud_dt", "Sede"], ascending=[True, True], inplace=True)

    tabs = st.tabs([
        "Resumen Ejecutivo",
        "Vacaciones Activas",
        "Responder Solicitudes",
        "Trazabilidad",
    ])

    with tabs[0]:
        top_cols = st.columns([1.55, 1])
        with top_cols[0]:
            risk_rows = pending_focus_df[
                [
                    "Solicitud_ID",
                    "Fecha_Solicitud",
                    "Nombre_Completo",
                    "Sede",
                    "Tipo_Solicitud",
                    "Estado",
                ]
            ].head(12)
            render_info_panel(
                "Casos que requieren atencion",
                "Solicitudes pendientes o en revision, ordenadas por antiguedad dentro del corte activo.",
                "<div class='fx-mini-grid'>"
                f"<div class='fx-mini-card'><div class='mini-label'>Pendientes</div><div class='mini-value'>{pending_count}</div></div>"
                f"<div class='fx-mini-card'><div class='mini-label'>En revision</div><div class='mini-value'>{review_count}</div></div>"
                f"<div class='fx-mini-card'><div class='mini-label'>Negadas</div><div class='mini-value'>{denied_count}</div></div>"
                f"<div class='fx-mini-card'><div class='mini-label'>Exportacion</div><div class='mini-value'>Excel profesional listo</div></div>"
                "</div>",
            )
            st.write("")
            st.dataframe(risk_rows, use_container_width=True, hide_index=True)

        with top_cols[1]:
            top_sede = sede_summary.iloc[0]["Sede"] if not sede_summary.empty else "Sin dato"
            top_type = type_summary.iloc[0]["Tipo_Solicitud"] if not type_summary.empty else "Sin dato"
            top_employee = employee_summary.iloc[0]["Nombre_Completo"] if not employee_summary.empty else "Sin dato"
            executive_html = f"""
            <div class='fx-mini-grid'>
                <div class='fx-mini-card'>
                    <div class='mini-label'>Sede con mayor volumen</div>
                    <div class='mini-value'>{top_sede}</div>
                </div>
                <div class='fx-mini-card'>
                    <div class='mini-label'>Tipo predominante</div>
                    <div class='mini-value'>{top_type}</div>
                </div>
                <div class='fx-mini-card'>
                    <div class='mini-label'>Mayor recurrencia</div>
                    <div class='mini-value'>{top_employee}</div>
                </div>
                <div class='fx-mini-card'>
                    <div class='mini-label'>Corte aplicado</div>
                    <div class='mini-value'>{filter_summary['Rango de fechas']}</div>
                </div>
            </div>
            """
            render_info_panel(
                "Lectura ejecutiva del corte",
                "Datos sinteticos para direccion y talento humano sobre el comportamiento del periodo seleccionado.",
                executive_html,
            )
            st.write("")
            with st.expander("Ver consolidado de estados", expanded=False):
                st.dataframe(status_summary, use_container_width=True, hide_index=True)

    with tabs[1]:
        analytic_top = st.columns([1.25, 1])
        with analytic_top[0]:
            render_info_panel(
                "Monitoreo central de vacaciones",
                "Control vigente de colaboradores en vacaciones para seguimiento transversal de todas las sedes.",
                f"<div class='fx-mini-grid'><div class='fx-mini-card'><div class='mini-label'>Vacaciones activas hoy</div><div class='mini-value'>{len(active_vacations_df)}</div></div><div class='fx-mini-card'><div class='mini-label'>Programadas mas adelante</div><div class='mini-value'>{len(upcoming_vacations_df)}</div></div><div class='fx-mini-card'><div class='mini-label'>Fecha de corte Colombia</div><div class='mini-value'>{today_colombia.strftime('%d/%m/%Y')}</div></div></div>",
            )
        with analytic_top[1]:
            render_info_panel(
                "Visibilidad por sedes",
                "Identifique rapidamente donde hay mayor concentracion de vacaciones activas para balance operativo.",
                "<div class='fx-mini-grid'><div class='fx-mini-card'><div class='mini-label'>Sedes con vacaciones activas</div><div class='mini-value'>{}</div></div><div class='fx-mini-card'><div class='mini-label'>Estado monitoreado</div><div class='mini-value'>Aprobadas vigentes</div></div></div>".format(len(vacations_by_sede)),
            )

        chart_cols = st.columns(2)
        with chart_cols[0]:
            st.markdown("#### Vacaciones activas por sede")
            if not vacations_by_sede.empty:
                st.bar_chart(vacations_by_sede.set_index("Sede")["Cantidad"], use_container_width=True)
            else:
                st.caption("No hay vacaciones activas en la fecha actual de Colombia.")
        with chart_cols[1]:
            st.markdown("#### Próximas vacaciones aprobadas")
            next_vacations = upcoming_vacations_df[["Nombre_Completo", "Sede", "Fecha_Inicial", "Fecha_Final"]].head(10) if not upcoming_vacations_df.empty else pd.DataFrame()
            if next_vacations.empty:
                st.caption("No hay vacaciones futuras dentro del corte filtrado.")
            else:
                st.dataframe(next_vacations, use_container_width=True, hide_index=True)

        st.markdown("#### Personal actualmente en vacaciones")
        if active_vacations_df.empty:
            st.info("Hoy no hay colaboradores en vacaciones dentro del corte filtrado.")
        else:
            st.dataframe(
                active_vacations_df[[
                    "Nombre_Completo",
                    "Cedula",
                    "Cargo",
                    "Sede",
                    "Fecha_Inicial",
                    "Fecha_Final",
                    "Dias_Aprobados",
                    "Responsable_Revision",
                ]],
                use_container_width=True,
                hide_index=True,
            )

        with st.expander("Ver detalle adicional de vacaciones y sedes", expanded=False):
            if not vacations_by_sede.empty:
                st.dataframe(vacations_by_sede, use_container_width=True, hide_index=True)
            if not upcoming_vacations_df.empty:
                st.dataframe(
                    upcoming_vacations_df[["Nombre_Completo", "Sede", "Fecha_Inicial", "Fecha_Final", "Dias_Aprobados"]],
                    use_container_width=True,
                    hide_index=True,
                )

    selectable = filtered_df[["Solicitud_ID", "Nombre_Completo", "Estado", "Tipo_Solicitud", "Fecha_Solicitud"]].copy()
    selectable["Etiqueta"] = selectable.apply(
        lambda row: f"{row['Solicitud_ID']} | {row['Nombre_Completo']} | {row['Tipo_Solicitud']} | {row['Estado']} | {row['Fecha_Solicitud']}",
        axis=1,
    )
    selected_label = st.selectbox(
        "Solicitud activa para gestion y trazabilidad",
        selectable["Etiqueta"].tolist(),
        key="selected_request_label",
    )
    selected_id = selectable.loc[selectable["Etiqueta"] == selected_label, "Solicitud_ID"].iloc[0]
    selected_record = filtered_df[filtered_df["Solicitud_ID"] == selected_id].iloc[0].to_dict()

    with tabs[2]:
        top_manage_cols = st.columns([1.25, 1])
        with top_manage_cols[0]:
            render_info_panel(
                "Respuesta administrativa focalizada",
                "La gestion de aprobacion o negacion se separa de la analitica para reducir carga visual y acelerar la respuesta.",
                "<div class='fx-mini-grid'><div class='fx-mini-card'><div class='mini-label'>Solicitudes visibles</div><div class='mini-value'>{}</div></div><div class='fx-mini-card'><div class='mini-label'>Pendientes + revision</div><div class='mini-value'>{}</div></div></div>".format(total_requests, pending_count + review_count),
            )
        with top_manage_cols[1]:
            st.download_button(
                "Descargar Excel gerencial",
                data=report_excel,
                file_name=f"reporte_gerencial_solicitudes_{current_colombia_datetime().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

        st.markdown("#### Solicitudes del corte para respuesta")
        report_columns = [
            "Solicitud_ID",
            "Fecha_Solicitud",
            "Estado",
            "Tipo_Solicitud",
            "Motivo_Descripcion",
            "Nombre_Completo",
            "Cedula",
            "Sede",
            "Fecha_Inicial",
            "Fecha_Final",
            "Responsable_Revision",
        ]
        with st.expander("Ver listado consolidado del corte", expanded=False):
            st.dataframe(filtered_df[report_columns], use_container_width=True, hide_index=True)

        detail_html = """
        <div class='fx-mini-grid'>
            <div class='fx-mini-card'><div class='mini-label'>Empleado</div><div class='mini-value'>{Nombre_Completo}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Cedula</div><div class='mini-value'>{Cedula}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Cargo</div><div class='mini-value'>{Cargo}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Sede</div><div class='mini-value'>{Sede}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Tipo</div><div class='mini-value'>{Tipo_Solicitud}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Motivo</div><div class='mini-value'>{Motivo_Descripcion}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Fecha inicial</div><div class='mini-value'>{Fecha_Inicial}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Fecha final</div><div class='mini-value'>{Fecha_Final}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Hora salida</div><div class='mini-value'>{Hora_Salida}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Tiempo total</div><div class='mini-value'>{Tiempo_Total}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Persona a cargo</div><div class='mini-value'>{Persona_A_Cargo}</div></div>
            <div class='fx-mini-card'><div class='mini-label'>Estado actual</div><div class='mini-value'>{Estado}</div></div>
        </div>
        """.format(**{key: str(selected_record.get(key, "") or "No registrado") for key in [
            "Nombre_Completo",
            "Cedula",
            "Cargo",
            "Sede",
            "Tipo_Solicitud",
            "Motivo_Descripcion",
            "Fecha_Inicial",
            "Fecha_Final",
            "Hora_Salida",
            "Tiempo_Total",
            "Persona_A_Cargo",
            "Estado",
        ]})
        render_info_panel(
            "Ficha integral de la solicitud seleccionada",
            "Lectura ejecutiva para decidir aprobacion, negacion o seguimiento interno.",
            detail_html,
        )

        st.text_area(
            "Detalle registrado por el empleado",
            value=selected_record.get("Detalle_Solicitud", ""),
            disabled=True,
            height=120,
        )

        attachment_url = selected_record.get("Adjunto_URL", "")
        attachment_name = selected_record.get("Adjunto_Nombre", "")
        attachment_status = selected_record.get("Adjunto_Estado", "")
        attach_cols = st.columns([1, 2])
        if attachment_url:
            attach_cols[0].link_button(
                f"Abrir soporte: {attachment_name or 'adjunto'}",
                attachment_url,
                use_container_width=True,
            )
        elif attachment_status and attachment_status != "SIN_ADJUNTO":
            attach_cols[0].caption(f"Adjunto: {attachment_status}")
        attach_cols[1].caption("Los soportes se conservan como evidencia del proceso administrativo.")

        st.markdown("#### Respuesta administrativa")
        with st.form("gestion_respuesta_solicitud"):
            g1, g2, g3 = st.columns(3)
            estado = g1.selectbox(
                "Estado",
                ["Pendiente", "En revision", "Aprobada", "Negada"],
                index=["Pendiente", "En revision", "Aprobada", "Negada"].index(selected_record.get("Estado", "Pendiente")),
            )
            responsable = g2.text_input(
                "Responsable de revision",
                value=selected_record.get("Responsable_Revision", ""),
                placeholder="Nombre de quien responde",
            )
            medio = g3.selectbox(
                "Medio de respuesta",
                ["WhatsApp", "Correo", "WhatsApp y Correo", "Interno"],
                index=["WhatsApp", "Correo", "WhatsApp y Correo", "Interno"].index(
                    selected_record.get("Medio_Respuesta", "WhatsApp")
                    if selected_record.get("Medio_Respuesta", "") in ["WhatsApp", "Correo", "WhatsApp y Correo", "Interno"]
                    else "WhatsApp"
                ),
            )

            g4, g5, g6, g7 = st.columns(4)
            dias_aprobados = g4.text_input("Dias aprobados", value=selected_record.get("Dias_Aprobados", ""))
            periodo = g5.text_input("Periodo correspondiente", value=selected_record.get("Periodo_Correspondiente", ""))
            dias_pendientes = g6.text_input("Dias pendientes del periodo", value=selected_record.get("Dias_Pendientes_Periodo", ""))
            reincorporacion = g7.text_input("Fecha de reincorporacion", value=selected_record.get("Fecha_Reincorporacion", ""), placeholder="DD/MM/AAAA")

            observaciones = st.text_area(
                "Observaciones de talento humano",
                value=selected_record.get("Observaciones_RRHH", ""),
                height=120,
            )
            save_response = st.form_submit_button("Guardar respuesta administrativa", type="primary", use_container_width=True)

        if save_response:
            if not responsable.strip():
                st.error("Debe registrar el responsable de revision.")
            else:
                selected_record["Estado"] = estado
                selected_record["Responsable_Revision"] = responsable.strip()
                selected_record["Medio_Respuesta"] = medio
                selected_record["Dias_Aprobados"] = dias_aprobados.strip()
                selected_record["Periodo_Correspondiente"] = periodo.strip()
                selected_record["Dias_Pendientes_Periodo"] = dias_pendientes.strip()
                selected_record["Fecha_Reincorporacion"] = reincorporacion.strip()
                selected_record["Observaciones_RRHH"] = observaciones.strip()
                selected_record["Fecha_Respuesta"] = format_colombia_timestamp()
                selected_record["Whatsapp_Listo"] = "SI" if medio in ["WhatsApp", "WhatsApp y Correo"] else "NO"
                selected_record["Ultima_Actualizacion"] = format_colombia_timestamp()

                update_request_record(worksheets["registros"], selected_id, selected_record)
                append_audit_log(
                    worksheets["auditoria"],
                    selected_id,
                    "RESPUESTA_RRHH",
                    responsable.strip(),
                    f"Solicitud actualizada con estado {estado}.",
                )
                append_novedad(
                    worksheets["novedades"],
                    selected_record,
                    "RESPUESTA_RRHH",
                    observaciones.strip() or f"Solicitud marcada como {estado}.",
                    medio,
                    responsable.strip(),
                )
                st.success("La solicitud fue actualizada correctamente.")
                st.rerun()

        action_cols = st.columns(3)
        whatsapp_url = build_whatsapp_url(selected_record)
        if whatsapp_url:
            action_cols[0].link_button("Abrir respuesta por WhatsApp", whatsapp_url, use_container_width=True)
        else:
            action_cols[0].button("Abrir respuesta por WhatsApp", disabled=True, use_container_width=True)

        if action_cols[1].button("Enviar respuesta por correo", use_container_width=True):
            email_ok, email_message = send_employee_response_email(selected_record)
            if email_ok:
                selected_record["Correo_Respuesta_Empleado"] = "SI"
                selected_record["Ultima_Actualizacion"] = format_colombia_timestamp()
                update_request_record(worksheets["registros"], selected_id, selected_record)
                append_audit_log(
                    worksheets["auditoria"],
                    selected_id,
                    "EMAIL_RESPUESTA",
                    selected_record.get("Responsable_Revision", "Sistema"),
                    "Respuesta enviada por correo al empleado.",
                )
                st.success("Correo de respuesta enviado al empleado.")
                st.rerun()
            st.error(f"No se pudo enviar el correo: {email_message}")

        if action_cols[2].button("Marcar en revision", use_container_width=True):
            selected_record["Estado"] = "En revision"
            selected_record["Ultima_Actualizacion"] = format_colombia_timestamp()
            update_request_record(worksheets["registros"], selected_id, selected_record)
            append_audit_log(
                worksheets["auditoria"],
                selected_id,
                "EN_REVISION",
                "Sistema",
                "Solicitud marcada manualmente en revision.",
            )
            st.success("Estado actualizado a En revision.")
            st.rerun()

    with tabs[3]:
        selected_novedades = novedades_df[novedades_df["Solicitud_ID"] == selected_id]
        selected_audit = audit_df[audit_df["Solicitud_ID"] == selected_id]
        trace_tabs = st.tabs(["Novedades de la solicitud", "Auditoria de la solicitud", "Trazabilidad global reciente"])
        with trace_tabs[0]:
            if selected_novedades.empty:
                st.caption("Sin novedades registradas para la solicitud seleccionada.")
            else:
                st.dataframe(selected_novedades, use_container_width=True, hide_index=True)
        with trace_tabs[1]:
            if selected_audit.empty:
                st.caption("Sin eventos de auditoria para la solicitud seleccionada.")
            else:
                st.dataframe(selected_audit, use_container_width=True, hide_index=True)
        with trace_tabs[2]:
            global_novedades = novedades_df.sort_values(by="Fecha_Evento", ascending=False).head(25) if not novedades_df.empty else novedades_df
            global_audit = audit_df.sort_values(by="Fecha_Evento", ascending=False).head(25) if not audit_df.empty else audit_df
            st.markdown("##### Ultimas novedades registradas")
            if global_novedades.empty:
                st.caption("Sin novedades globales registradas.")
            else:
                st.dataframe(global_novedades, use_container_width=True, hide_index=True)
            st.markdown("##### Ultimos eventos de auditoria")
            if global_audit.empty:
                st.caption("Sin auditoria global registrada.")
            else:
                st.dataframe(global_audit, use_container_width=True, hide_index=True)


if __name__ == "__main__":
    main()
